import unittest
import io
import openpyxl
from fastapi.testclient import TestClient

from api.index import app, generate_full_excel_report


class TestExcelExport(unittest.TestCase):
    def setUp(self):
        self.client = TestClient(app)

    def test_get_export_full_excel(self):
        """Test GET /api/export-full-excel endpoint returns valid multi-sheet Excel file."""
        response = self.client.get("/api/export-full-excel?province=sulawesi_selatan&compare_province=gorontalo&year=2024&analysis_mode=growth_yoy&transport_type=penumpang")
        self.assertEqual(response.status_code, 200)
        self.assertIn("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", response.headers["content-type"])
        self.assertIn("attachment; filename=", response.headers.get("content-disposition", ""))

        # Load workbook from response bytes
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        expected_sheets = ["1_Korelasi_PDRB", "2_Diagnostik_VIF", "3_Komparasi_Wilayah", "4_Tren_Triwulanan", "5_Data_Mentah"]
        self.assertEqual(wb.sheetnames, expected_sheets)

        # Check Sheet 1
        ws1 = wb["1_Korelasi_PDRB"]
        self.assertGreater(ws1.max_row, 10)
        self.assertEqual(ws1.max_column, 12)

        # Check Sheet 2 (Diagnostik VIF, Korelasi Silang X x Y, and Multiko Transport Y x Y)
        ws2 = wb["2_Diagnostik_VIF"]
        self.assertGreaterEqual(ws2.max_row, 35)

        cell_texts = [str(ws2.cell(row=r, column=1).value or "") for r in range(1, ws2.max_row + 1)]
        all_text = " ".join(cell_texts)
        self.assertIn("BAGIAN A: DIAGNOSTIK MULTIKOLINEARITAS", all_text)
        self.assertIn("BAGIAN B: KORELASI SILANG PDRB", all_text)
        self.assertIn("BAGIAN C: MULTIKOLINEARITAS ANTAR-INDIKATOR TRANSPORTASI", all_text)

        # Check Sheet 3
        ws3 = wb["3_Komparasi_Wilayah"]
        self.assertGreater(ws3.max_row, 10)
        self.assertEqual(ws3.max_column, 9)

        # Check Sheet 4
        ws4 = wb["4_Tren_Triwulanan"]
        self.assertGreater(ws4.max_row, 10)
        self.assertEqual(ws4.max_column, 15)

        # Check Sheet 5
        ws5 = wb["5_Data_Mentah"]
        self.assertGreater(ws5.max_row, 10)

    def test_post_export_full_excel(self):
        """Test POST /api/export-full-excel endpoint with JSON payload."""
        payload = {
            "province": "kalimantan_timur",
            "compare_province": "kalimantan_barat",
            "year": "2023",
            "analysis_mode": "abs_hk",
            "transport_type": "barang"
        }
        response = self.client.post("/api/export-full-excel", json=payload)
        self.assertEqual(response.status_code, 200)

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        self.assertEqual(len(wb.sheetnames), 5)


if __name__ == "__main__":
    unittest.main()
