"""
========================================================================================
FastAPI Serverless Application - Korelasi PDRB & Transportasi (Vercel Serverless Ready)
Dynamic Path Resolution, In-Memory Parquet Cache, dan Full Diagnostic Logging.
v4.0 — Growth Rate Analysis (YoY/QoQ) + Pooled Multi-Year Correlation
========================================================================================
"""

import io
import os
import sys
import json
import re
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple

from fastapi import FastAPI, Request, Query, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
import pandas as pd
import numpy as np
from scipy.stats import pearsonr
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ======================================================================================
# DYNAMIC PATH RESOLUTION FOR VERCEL LINUX SERVERLESS & LOCAL ENVIRONMENT
# ======================================================================================

def resolve_data_dir() -> Path:
    """Mencari lokasi folder data/ secara dinamis di berbagai lingkungan Vercel & Lokal."""
    candidates = [
        Path(__file__).resolve().parent / "data",
        Path(__file__).resolve().parent.parent / "data",
        Path(os.getcwd()) / "data",
        Path(os.getcwd()) / "api" / "data",
        Path("/var/task/data"),
        Path("/var/task/api/data"),
    ]
    for p in candidates:
        if p.exists() and (p / "manifest.json").exists():
            return p
    for p in candidates:
        if p.exists() and len(list(p.glob("*.parquet"))) > 0:
            return p
    return Path(__file__).resolve().parent / "data"


def resolve_templates_dir() -> Path:
    """Mencari lokasi folder templates/ secara dinamis."""
    candidates = [
        Path(__file__).resolve().parent.parent / "templates",
        Path(__file__).resolve().parent / "templates",
        Path(os.getcwd()) / "templates",
        Path(os.getcwd()) / "api" / "templates",
        Path("/var/task/templates"),
        Path("/var/task/api/templates"),
    ]
    for p in candidates:
        if p.exists() and (p / "index.html").exists():
            return p
    return Path(__file__).resolve().parent.parent / "templates"


def resolve_public_dir() -> Path:
    """Mencari lokasi folder public/ secara dinamis."""
    candidates = [
        Path(__file__).resolve().parent.parent / "public",
        Path(__file__).resolve().parent / "public",
        Path(os.getcwd()) / "public",
        Path("/var/task/public"),
    ]
    for p in candidates:
        if p.exists():
            return p
    return Path(__file__).resolve().parent.parent / "public"


DATA_DIR = resolve_data_dir()
TEMPLATES_DIR = resolve_templates_dir()
PUBLIC_DIR = resolve_public_dir()

# Inisialisasi FastAPI App
app = FastAPI(
    title="Korelasi PDRB & Transportasi API",
    description="Serverless API & Dashboard Analisis PDRB dan Transportasi BPS",
    version="4.0.0"
)

# Mount static files jika direktori ada
if PUBLIC_DIR.exists():
    try:
        app.mount("/public", StaticFiles(directory=str(PUBLIC_DIR)), name="public")
    except Exception as e:
        print(f"⚠️ Warning mounting /public: {e}")

# Setup Jinja2 Templates
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

import functools

# Cache data parquet di memory
_parquet_cache: Dict[str, pd.DataFrame] = {}

# Cache multi-year aggregated DataFrames
_multi_year_cache: Dict[str, pd.DataFrame] = {}

# Cache headers standar untuk Vercel Edge CDN (tanpa browser max-age agar selalu revalidate)
CACHE_HEADERS = {
    "Cache-Control": "public, s-maxage=86400, stale-while-revalidate=86400"
}


def cached_json_response(content: Any) -> JSONResponse:
    """Mengembalikan JSONResponse dengan HTTP Cache-Control headers terstandar."""
    return JSONResponse(content=content, headers=CACHE_HEADERS)


# Mapping triwulan string → integer yang aman (case-insensitive, toleran terhadap variasi)
_TRIWULAN_MAP: Dict[str, int] = {
    "triwulan i": 1, "triwulan 1": 1, "q1": 1, "tw i": 1, "tw 1": 1, "i": 1,
    "triwulan ii": 2, "triwulan 2": 2, "q2": 2, "tw ii": 2, "tw 2": 2, "ii": 2,
    "triwulan iii": 3, "triwulan 3": 3, "q3": 3, "tw iii": 3, "tw 3": 3, "iii": 3,
    "triwulan iv": 4, "triwulan 4": 4, "q4": 4, "tw iv": 4, "tw 4": 4, "iv": 4,
}


def _parse_triwulan_num(label: str) -> int:
    """Konversi label triwulan ke integer 1-4 secara aman."""
    key = str(label).strip().lower()
    if key in _TRIWULAN_MAP:
        return _TRIWULAN_MAP[key]
    # Fallback: cari angka romawi atau digit di dalam string
    roman_match = re.search(r'\b(iv|iii|ii|i)\b', key)
    if roman_match:
        roman_to_int = {"i": 1, "ii": 2, "iii": 3, "iv": 4}
        return roman_to_int.get(roman_match.group(1), 1)
    digit_match = re.search(r'(\d)', key)
    if digit_match:
        d = int(digit_match.group(1))
        return d if 1 <= d <= 4 else 1
    return 1


@functools.lru_cache(maxsize=1)
def load_manifest() -> Dict[str, Any]:
    """Membaca manifest metadata data parquet (memoized in-memory)."""
    manifest_path = DATA_DIR / "manifest.json"
    if manifest_path.exists():
        try:
            with open(manifest_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"⚠️ Error reading manifest: {e}")

    # Fallback auto-discovery
    fallback_provinces: Dict[str, Any] = {}
    if DATA_DIR.exists():
        for p_file in DATA_DIR.glob("*.parquet"):
            year_match = re.search(r"20\d{2}", p_file.stem)
            if year_match:
                y = year_match.group(0)
                prov_key = p_file.stem[:p_file.stem.find(y)-1]
                prov_name = " ".join([w.capitalize() for w in prov_key.split("_")])
                if prov_key not in fallback_provinces:
                    fallback_provinces[prov_key] = {"name": prov_name, "years": {}}
                if y not in fallback_provinces[prov_key]["years"]:
                    fallback_provinces[prov_key]["years"][y] = {"sheets": []}
                sheet_name = p_file.stem[p_file.stem.find(y)+5:]
                fallback_provinces[prov_key]["years"][y]["sheets"].append(sheet_name)

    return {"provinces": fallback_provinces, "auto_discovered": True}


@functools.lru_cache(maxsize=256)
def _read_parquet_cached_file(fp_str: str, cols_tuple: Optional[Tuple[str, ...]] = None) -> Optional[pd.DataFrame]:
    """Membaca file parquet dari disk dengan LRU cache."""
    try:
        cols = list(cols_tuple) if cols_tuple else None
        return pd.read_parquet(fp_str, columns=cols)
    except Exception as e:
        print(f"❌ Error reading parquet {fp_str}: {e}")
        return None


def get_parquet_df(file_stem: str, columns: Optional[List[str]] = None) -> Optional[pd.DataFrame]:
    """Membaca file parquet dengan in-memory cache, multi-candidate path search, dan column pruning."""
    cache_key = f"{file_stem}_{','.join(sorted(columns)) if columns else 'all'}"
    if cache_key in _parquet_cache:
        return _parquet_cache[cache_key]

    cols_tuple = tuple(sorted(columns)) if columns else None
    candidate_paths = [
        DATA_DIR / f"{file_stem}.parquet",
        resolve_data_dir() / f"{file_stem}.parquet",
        Path(__file__).resolve().parent / "data" / f"{file_stem}.parquet",
        Path(__file__).resolve().parent.parent / "data" / f"{file_stem}.parquet",
    ]

    for fp in candidate_paths:
        if fp.exists():
            df = _read_parquet_cached_file(str(fp), cols_tuple)
            if df is not None:
                _parquet_cache[cache_key] = df
                return df

    return None


# ======================================================================================
# GROWTH RATE ANALYSIS ENGINE — Multi-Year Aggregation & Pertumbuhan (Memoized)
# ======================================================================================

@functools.lru_cache(maxsize=32)
def _build_multi_year_df_cached(province: str) -> Tuple[pd.DataFrame, Tuple[str, ...]]:
    """
    Menggabungkan semua file pdrb_triwulan.parquet lintas tahun untuk satu provinsi
    menjadi satu DataFrame panjang yang tersortir kronologis (Memoized).
    """
    manifest = load_manifest()
    prov_info = manifest.get("provinces", {}).get(province, {})
    available_years = sorted(list(prov_info.get("years", {}).keys())) if prov_info.get("years") else []

    # Fallback: scan data directory
    if not available_years:
        for fp in DATA_DIR.glob(f"{province}_*_pdrb_triwulan.parquet"):
            ym = re.search(r"(\d{4})", fp.stem)
            if ym:
                available_years.append(ym.group(1))
        available_years = sorted(set(available_years))

    frames = []
    for yr in available_years:
        df_yr = get_parquet_df(f"{province}_{yr}_pdrb_triwulan")
        if df_yr is not None and "Triwulan" in df_yr.columns:
            df_copy = df_yr.copy()
            df_copy["_tahun"] = int(yr)
            df_copy["_triwulan_num"] = df_copy["Triwulan"].apply(_parse_triwulan_num)
            # Kolom sorting kronologis: tahun * 10 + triwulan_num
            df_copy["_time_order"] = df_copy["_tahun"] * 10 + df_copy["_triwulan_num"]
            frames.append(df_copy)

    if not frames:
        return pd.DataFrame(), tuple(available_years)

    df_combined = pd.concat(frames, ignore_index=True)
    df_combined = df_combined.sort_values("_time_order").reset_index(drop=True)
    return df_combined, tuple(available_years)


def build_multi_year_df(province: str) -> Tuple[pd.DataFrame, List[str]]:
    """Public wrapper untuk build_multi_year_df dengan return type DataFrame & list."""
    df, years = _build_multi_year_df_cached(province)
    return df.copy(), list(years)


@functools.lru_cache(maxsize=64)
def _get_growth_df_cached(province: str, periods: int) -> pd.DataFrame:
    """Menghitung dan mem-cache laju pertumbuhan multi-tahun per provinsi & period (YoY/QoQ)."""
    df_multi, _ = _build_multi_year_df_cached(province)
    if df_multi.empty:
        return pd.DataFrame()

    df_g = df_multi.copy()
    numeric_cols = [c for c in df_g.columns
                    if c not in ("Triwulan", "_tahun", "_triwulan_num", "_time_order")
                    and pd.api.types.is_numeric_dtype(df_g[c])]

    for col in numeric_cols:
        df_g[col] = df_g[col].pct_change(periods=periods) * 100.0

    df_g.replace([np.inf, -np.inf], np.nan, inplace=True)
    df_g = df_g.dropna(subset=numeric_cols, how="all").reset_index(drop=True)
    return df_g


def compute_growth_df(df: pd.DataFrame, periods: int = 4) -> pd.DataFrame:
    """
    Menghitung laju pertumbuhan (pct_change) pada seluruh kolom numerik.
    """
    if df.empty:
        return df

    df_g = df.copy()
    numeric_cols = [c for c in df_g.columns
                    if c not in ("Triwulan", "_tahun", "_triwulan_num", "_time_order")
                    and pd.api.types.is_numeric_dtype(df_g[c])]

    for col in numeric_cols:
        df_g[col] = df_g[col].pct_change(periods=periods) * 100.0

    df_g.replace([np.inf, -np.inf], np.nan, inplace=True)
    df_g = df_g.dropna(subset=numeric_cols, how="all").reset_index(drop=True)
    return df_g


def _compute_correlation_matrix_from_growth(
    df_growth: pd.DataFrame,
    category: str = "lu",
    pdrb_price_type: str = "HK"
) -> pd.DataFrame:
    """
    Menghitung matriks korelasi (+ p-value) antara setiap sektor PDRB
    dan 3 metrik transportasi dari data pertumbuhan.
    Mendukung pdrb_price_type: 'HK', 'HB', atau 'ALL'/'all'.
    """
    if str(pdrb_price_type).upper() == "ALL":
        df_hk = _compute_correlation_matrix_from_growth(df_growth, category=category, pdrb_price_type="HK")
        df_hb = _compute_correlation_matrix_from_growth(df_growth, category=category, pdrb_price_type="HB")
        if df_hk.empty:
            return df_hb
        if df_hb.empty:
            return df_hk

        label_col = "Lapangan Usaha" if category == "lu" else "Komponen Pengeluaran"
        combined_rows = []
        hk_rows = df_hk.to_dict(orient="records")
        hb_rows = df_hb.to_dict(orient="records")
        hb_by_label = {r[label_col]: r for r in hb_rows if label_col in r}
        used_hb = set()

        for r_hk in hk_rows:
            combined_rows.append(r_hk)
            lbl = r_hk.get(label_col)
            if lbl in hb_by_label:
                combined_rows.append(hb_by_label[lbl])
                used_hb.add(lbl)

        for r_hb in hb_rows:
            if r_hb.get(label_col) not in used_hb:
                combined_rows.append(r_hb)

        return pd.DataFrame(combined_rows)

    p_col = next((c for c in df_growth.columns if "penumpang" in c.lower()), None)
    bag_col = next((c for c in df_growth.columns if "bagasi" in c.lower()), None)
    bar_col = next((c for c in df_growth.columns if "barang" in c.lower()), None)

    is_hb = str(pdrb_price_type).upper() == "HB"
    prefix = ("LU (HB)" if is_hb else "LU (HK)") if category == "lu" else ("Peng (HB)" if is_hb else "Peng (HK)")
    label_col = "Lapangan Usaha" if category == "lu" else "Komponen Pengeluaran"
    tipe_label = "Harga Berlaku (HB)" if is_hb else "Harga Konstan (HK)"

    sector_cols = [c for c in df_growth.columns if c.lower().startswith(prefix.lower())]

    rows = []
    for sec_col in sector_cols:
        clean_name = re.sub(r"^(LU|Peng)\s*\([A-Z]+\)\s*-\s*", "", sec_col).strip()
        row: Dict[str, Any] = {label_col: clean_name, "Tipe PDRB": tipe_label}

        for metric_label, metric_col in [("Penumpang", p_col), ("Bagasi", bag_col), ("Barang", bar_col)]:
            cor_key = f"Korelasi dgn {metric_label}"
            pv_key = f"p-value {metric_label}"

            if metric_col is None or metric_col not in df_growth.columns:
                row[cor_key] = None
                row[pv_key] = None
                continue

            mask = df_growth[[sec_col, metric_col]].dropna()
            x_vals = mask[sec_col].values.astype(float)
            y_vals = mask[metric_col].values.astype(float)

            if len(x_vals) >= 3:
                try:
                    r_val, p_val = pearsonr(x_vals, y_vals)
                    row[cor_key] = float(r_val) if not np.isnan(r_val) else None
                    row[pv_key] = float(p_val) if not np.isnan(p_val) else None
                except Exception:
                    row[cor_key] = None
                    row[pv_key] = None
            else:
                row[cor_key] = None
                row[pv_key] = None

        rows.append(row)

    return pd.DataFrame(rows)


@functools.lru_cache(maxsize=128)
def _get_growth_correlation_matrix_cached(
    province: str,
    periods: int,
    category: str,
    pdrb_price_type: str
) -> pd.DataFrame:
    """Memoized base correlation matrix for growth mode."""
    df_growth = _get_growth_df_cached(province, periods)
    if df_growth.empty:
        return pd.DataFrame()
    return _compute_correlation_matrix_from_growth(df_growth, category=category, pdrb_price_type=pdrb_price_type)


# ======================================================================================
# API ENDPOINTS (with Caching & High Performance)
# ======================================================================================

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    import traceback
    err_trace = traceback.format_exc()
    print(f"⚠️ Unhandled Server Error at {request.url.path}: {exc}\n{err_trace}")
    return JSONResponse(
        status_code=500,
        content={
            "status": "error",
            "error": str(exc),
            "type": exc.__class__.__name__,
            "path": request.url.path,
            "message": "Terjadi kesalahan internal pada serverless engine. Silakan muat ulang."
        }
    )


@app.get("/", response_class=HTMLResponse)
async def index_page(request: Request):
    """Serve halaman dashboard utama."""
    manifest = load_manifest()
    manifest_json = "{}"
    try:
        manifest_json = json.dumps(manifest)
    except Exception as e:
        print(f"⚠️ Error dumping manifest to JSON: {e}")

    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={
            "title": "Dashboard Analisis PDRB & Transportasi",
            "manifest": manifest,
            "manifest_json": manifest_json,
            "data_dir": str(DATA_DIR)
        }
    )


@app.get("/api/health")
@app.get("/api/warmup")
async def health_check():
    """Health check & pre-warm container serverless."""
    data_files = [f.name for f in DATA_DIR.glob("*")] if DATA_DIR.exists() else []
    return cached_json_response({
        "status": "ok",
        "service": "KorelasiPDRB-FastAPI",
        "version": "4.0.0",
        "data_dir": str(DATA_DIR),
        "data_dir_exists": DATA_DIR.exists(),
        "total_data_files": len(data_files),
        "sample_files": data_files[:5]
    })


@app.get("/api/manifest")
async def get_manifest():
    """Mengembalikan seluruh manifest metadata."""
    return cached_json_response(load_manifest())


@functools.lru_cache(maxsize=64)
def _compute_options(province: str, year: str) -> Dict[str, Any]:
    manifest = load_manifest()
    lu_stem = f"{province}_{year}_pdrb_correl_lu"
    df_lu = get_parquet_df(lu_stem)
    
    sectors = []
    if df_lu is not None and "Lapangan Usaha" in df_lu.columns:
        sectors = sorted(df_lu["Lapangan Usaha"].dropna().unique().tolist())
    
    return {
        "manifest": manifest,
        "selected_province": province,
        "selected_year": year,
        "sectors": sectors,
        "transport_metrics": [
            {"key": "penumpang", "label": "✈️ Penumpang (Orang)"},
            {"key": "bagasi", "label": "🧳 Bagasi (Kg)"},
            {"key": "barang", "label": "📦 Barang/Kargo (Kg)"}
        ]
    }


@app.get("/api/options")
async def get_options(province: str = "sulawesi_selatan", year: str = "2024"):
    """Mengembalikan opsi filter (provinsi, tahun, sektor, dan tipe PDRB)."""
    return cached_json_response(_compute_options(province, year))


@functools.lru_cache(maxsize=512)
def _compute_correlations(
    province: str,
    year: str,
    type: str,
    filter_mode: str,
    sort_by: str,
    pdrb_type: str,
    search: str,
    analysis_mode: str
) -> Dict[str, Any]:
    is_growth = analysis_mode.startswith("growth_")
    label_col = "Lapangan Usaha" if type == "lu" else "Komponen Pengeluaran"
    data_warning = None

    if is_growth:
        periods = 4 if "yoy" in analysis_mode else 1
        if analysis_mode.endswith("_hb"):
            growth_pdrb_type = "HB"
        elif analysis_mode.endswith("_all") or str(pdrb_type).lower() == "all":
            growth_pdrb_type = "ALL"
        elif str(pdrb_type).upper() == "HB":
            growth_pdrb_type = "HB"
        else:
            growth_pdrb_type = "HK"

        df_multi, avail_years = _build_multi_year_df_cached(province)

        if df_multi.empty:
            return {
                "type": type,
                "analysis_mode": analysis_mode,
                "count": 0,
                "rows": [],
                "error": f"Data triwulan multi-tahun untuk {province} tidak ditemukan."
            }

        total_quarters = len(df_multi)
        if "yoy" in analysis_mode and total_quarters <= 4:
            data_warning = (
                f"Data hanya tersedia {total_quarters} triwulan. "
                f"YoY membutuhkan >4 triwulan. Pertimbangkan gunakan mode QoQ."
            )

        if "yoy" in analysis_mode and avail_years and str(year) == str(avail_years[0]):
            data_warning = (
                f"Tahun {year} adalah tahun pertama dalam dataset. "
                f"Tidak ada data periode sebelumnya untuk menghitung YoY tahun tunggal. "
                f"Korelasi dihitung dari seluruh tahun yang tersedia (pooled multi-tahun)."
            )

        df_growth = _get_growth_df_cached(province, periods)

        if df_growth.empty:
            return {
                "type": type,
                "analysis_mode": analysis_mode,
                "count": 0,
                "rows": [],
                "data_warning": data_warning,
                "error": "Data pertumbuhan kosong setelah kalkulasi (kemungkinan rentang tahun terlalu pendek)."
            }

        df_view = _get_growth_correlation_matrix_cached(province, periods, type, growth_pdrb_type).copy()

        if growth_pdrb_type == "HB":
            growth_pdrb_label = "Harga Berlaku (Nominal)"
        elif growth_pdrb_type == "ALL":
            growth_pdrb_label = "Semua Tipe (HK & HB)"
        else:
            growth_pdrb_label = "Harga Konstan (Riil)"

        analysis_label = f"Pertumbuhan {'YoY' if periods == 4 else 'QoQ'} (%) — {growth_pdrb_label}"
        n_observations = len(df_growth)

    else:
        stem = f"{province}_{year}_pdrb_correl_lu" if type == "lu" else f"{province}_{year}_pdrb_correl_peng"
        df = get_parquet_df(stem)

        if df is None:
            return {
                "type": type,
                "analysis_mode": analysis_mode,
                "count": 0,
                "rows": [],
                "error": f"Data korelasi {stem}.parquet tidak ditemukan di {DATA_DIR}."
            }

        df_view = df.copy()

        effective_pdrb_type = "HK" if analysis_mode == "abs_hk" else ("HB" if analysis_mode == "abs_hb" else ("all" if analysis_mode == "abs_all" else pdrb_type))
        if effective_pdrb_type == "HB" and "Tipe PDRB" in df_view.columns:
            df_view = df_view[df_view["Tipe PDRB"].str.contains("HB", case=False, na=False)]
        elif effective_pdrb_type == "HK" and "Tipe PDRB" in df_view.columns:
            df_view = df_view[df_view["Tipe PDRB"].str.contains("HK", case=False, na=False)]

        if effective_pdrb_type == "HB":
            analysis_label = "Nilai Nominal (Harga Berlaku)"
        elif effective_pdrb_type == "HK":
            analysis_label = "Nilai Riil (Harga Konstan)"
        else:
            analysis_label = "Semua Tipe (HK & HB)"
        n_observations = 4

    # Filters
    if search and label_col in df_view.columns:
        df_view = df_view[df_view[label_col].str.contains(search, case=False, na=False)]

    if filter_mode == "6_core" and type == "lu":
        core_keywords = [
            "perdagangan",
            "konstruksi",
            "industri pengolahan",
            "transportasi dan pergudangan",
            "akomodasi",
            "pengadaan air"
        ]
        pattern = "|".join(core_keywords)
        df_view = df_view[df_view[label_col].str.contains(pattern, case=False, na=False)]
    elif filter_mode == "top5_high" and "Korelasi dgn Penumpang" in df_view.columns:
        df_view = df_view.sort_values(by="Korelasi dgn Penumpang", ascending=False).head(10)
    elif filter_mode == "top5_low" and "Korelasi dgn Penumpang" in df_view.columns:
        df_view = df_view.sort_values(by="Korelasi dgn Penumpang", ascending=True).head(10)
    elif filter_mode == "strong_70":
        cor_cols = [c for c in df_view.columns if "Korelasi" in c]
        if cor_cols:
            cond = df_view[cor_cols].abs().ge(0.70).any(axis=1)
            df_view = df_view[cond]
    elif filter_mode == "transport_only" and label_col in df_view.columns:
        df_view = df_view[df_view[label_col].str.contains("transportasi", case=False, na=False)]

    # Sorting
    if sort_by == "p_desc" and "Korelasi dgn Penumpang" in df_view.columns:
        df_view = df_view.sort_values(by="Korelasi dgn Penumpang", ascending=False)
    elif sort_by == "p_asc" and "Korelasi dgn Penumpang" in df_view.columns:
        df_view = df_view.sort_values(by="Korelasi dgn Penumpang", ascending=True)
    elif sort_by == "bag_desc" and "Korelasi dgn Bagasi" in df_view.columns:
        df_view = df_view.sort_values(by="Korelasi dgn Bagasi", ascending=False)
    elif sort_by == "bag_asc" and "Korelasi dgn Bagasi" in df_view.columns:
        df_view = df_view.sort_values(by="Korelasi dgn Bagasi", ascending=True)
    elif sort_by == "bar_desc" and "Korelasi dgn Barang" in df_view.columns:
        df_view = df_view.sort_values(by="Korelasi dgn Barang", ascending=False)
    elif sort_by == "bar_asc" and "Korelasi dgn Barang" in df_view.columns:
        df_view = df_view.sort_values(by="Korelasi dgn Barang", ascending=True)

    df_view = df_view.replace({np.nan: None})
    rows = df_view.to_dict(orient="records")

    result: Dict[str, Any] = {
        "type": type,
        "label_col": label_col,
        "count": len(rows),
        "rows": rows,
        "analysis_mode": analysis_mode,
        "analysis_label": analysis_label,
        "n_observations": n_observations,
    }
    if data_warning:
        result["data_warning"] = data_warning

    return result


@app.get("/api/correlations")
async def get_correlations(
    province: str = "sulawesi_selatan",
    year: str = "2024",
    type: str = "lu",
    filter_mode: str = "all",
    sort_by: str = "default",
    pdrb_type: str = "all",
    search: str = "",
    analysis_mode: str = "growth_yoy"
):
    """Mengembalikan data matriks korelasi dengan in-memory cache berkecepatan tinggi."""
    res = _compute_correlations(province, year, type, filter_mode, sort_by, pdrb_type, search, analysis_mode)
    return cached_json_response(res)


@functools.lru_cache(maxsize=1024)
def _compute_dashboard_data(
    province: str,
    year: str,
    sector: str,
    category: str,
    tipe_pdrb: str,
    transport_metric: str,
    analysis_mode: str,
    ols_scope: str
) -> Dict[str, Any]:
    is_growth = analysis_mode.startswith("growth_")
    data_warning = None

    if is_growth:
        if analysis_mode.endswith("_hb"):
            chosen_type = "HB"
        elif analysis_mode.endswith("_all"):
            chosen_type = "HB" if str(tipe_pdrb).upper() == "HB" else "HK"
        else:
            chosen_type = "HB" if str(tipe_pdrb).upper() == "HB" else "HK"
    elif analysis_mode == "abs_hb":
        chosen_type = "HB"
    elif analysis_mode == "abs_hk":
        chosen_type = "HK"
    else:
        chosen_type = "HB" if str(tipe_pdrb).upper() == "HB" else "HK"

    stem_tri = f"{province}_{year}_pdrb_triwulan"
    df_tri = get_parquet_df(stem_tri)

    if df_tri is None:
        return {"error": f"Data triwulan {stem_tri}.parquet tidak ditemukan di {DATA_DIR}."}

    p_col = next((c for c in df_tri.columns if "penumpang" in c.lower()), None)
    bag_col = next((c for c in df_tri.columns if "bagasi" in c.lower()), None)
    bar_col = next((c for c in df_tri.columns if "barang" in c.lower()), None)

    lu_hk_cols = [c for c in df_tri.columns if c.startswith("LU (HK)")]
    lu_hb_cols = [c for c in df_tri.columns if c.startswith("LU (HB)")]
    total_lu_hk_col = next((c for c in lu_hk_cols if "produk domestik" in c.lower() or "pdrb" in c.lower()), None)

    clean_sec = re.sub(r'^(LU|Peng)\s*\([A-Z]+\)\s*-\s*', '', sector, flags=re.IGNORECASE).strip().lower()
    prefix_hk = "Peng (HK)" if category == "peng" else "LU (HK)"
    prefix_hb = "Peng (HB)" if category == "peng" else "LU (HB)"

    matching_hk = [c for c in df_tri.columns if clean_sec in c.lower() and prefix_hk.lower() in c.lower()]
    matching_hb = [c for c in df_tri.columns if clean_sec in c.lower() and prefix_hb.lower() in c.lower()]

    col_hk = matching_hk[0] if matching_hk else None
    col_hb = matching_hb[0] if matching_hb else None

    actual_tipe = chosen_type
    if chosen_type == "HB" and col_hb:
        sector_col = col_hb
        actual_tipe = "HB"
    elif col_hk:
        sector_col = col_hk
        actual_tipe = "HK"
    elif col_hb:
        sector_col = col_hb
        actual_tipe = "HB"
    else:
        sector_col = total_lu_hk_col or df_tri.columns[1]
        actual_tipe = "HK"

    active_label = re.sub(r'^(LU|Peng)\s*\([A-Z]+\)\s*-\s*', '', sector_col).strip()

    # 1. KPI Cards
    tot_p = float(df_tri[p_col].sum()) if p_col else 0
    tot_bag = float(df_tri[bag_col].sum()) if bag_col else 0
    tot_bar = float(df_tri[bar_col].sum()) if bar_col else 0
    tot_pdrb_hk = float(df_tri[total_lu_hk_col].sum()) if total_lu_hk_col else 0

    def get_q4_growth(col):
        if col and col in df_tri and len(df_tri) > 1:
            q_last = float(df_tri[col].iloc[-1])
            q_prev = float(df_tri[col].iloc[-2])
            return ((q_last - q_prev) / q_prev * 100) if q_prev != 0 else 0
        return 0.0

    kpi = {
        "penumpang_total": tot_p,
        "penumpang_q4_growth": get_q4_growth(p_col),
        "bagasi_total": tot_bag,
        "bagasi_q4_growth": get_q4_growth(bag_col),
        "barang_total": tot_bar,
        "barang_q4_growth": get_q4_growth(bar_col),
        "pdrb_hk_total": tot_pdrb_hk,
        "pdrb_hk_q4_growth": get_q4_growth(total_lu_hk_col),
        "max_correl_sector": "Transportasi dan Pergudangan",
        "max_correl_val": 0.811,
        "max_correl_type": "HK"
    }

    stem_lu = f"{province}_{year}_pdrb_correl_lu"
    df_lu = get_parquet_df(stem_lu)
    if df_lu is not None and "Korelasi dgn Penumpang" in df_lu.columns:
        df_f = df_lu[~df_lu["Lapangan Usaha"].str.contains("Produk Domestik|PDRB", case=False, na=False)]
        if not df_f.empty:
            top_row = df_f.sort_values(by="Korelasi dgn Penumpang", ascending=False).iloc[0]
            kpi["max_correl_sector"] = str(top_row["Lapangan Usaha"])
            kpi["max_correl_val"] = float(top_row["Korelasi dgn Penumpang"])
            kpi["max_correl_type"] = str(top_row.get("Tipe PDRB", ""))

    # 2. Time Series Data
    triwulan_labels = df_tri["Triwulan"].tolist()
    
    def calc_idx100(series):
        base = series.iloc[0]
        return [(float(v) / float(base) * 100.0) if base != 0 and pd.notna(v) else 100.0 for v in series]

    def calc_qoq(series):
        pct = series.pct_change() * 100.0
        return [0.0 if pd.isna(v) else float(v) for v in pct]

    pdrb_series = df_tri[sector_col].astype(float)
    pdrb_hk_series = df_tri[col_hk].astype(float) if col_hk else pdrb_series
    pdrb_hb_series = df_tri[col_hb].astype(float) if col_hb else pdrb_series

    p_series = df_tri[p_col].astype(float) if p_col else pd.Series([0]*len(df_tri))
    bag_series = df_tri[bag_col].astype(float) if bag_col else pd.Series([0]*len(df_tri))
    bar_series = df_tri[bar_col].astype(float) if bar_col else pd.Series([0]*len(df_tri))

    pdrb_idx = calc_idx100(pdrb_series)
    pdrb_qoq = calc_qoq(pdrb_series)
    pdrb_hk_idx = calc_idx100(pdrb_hk_series)
    pdrb_hk_qoq = calc_qoq(pdrb_hk_series)
    pdrb_hb_idx = calc_idx100(pdrb_hb_series)
    pdrb_hb_qoq = calc_qoq(pdrb_hb_series)

    p_idx = calc_idx100(p_series)
    p_qoq = calc_qoq(p_series)
    bag_idx = calc_idx100(bag_series)
    bag_qoq = calc_qoq(bag_series)
    bar_idx = calc_idx100(bar_series)
    bar_qoq = calc_qoq(bar_series)

    growth_ts_data: Dict[str, List[float]] = {}
    df_multi_all, avail_years_tuple = _build_multi_year_df_cached(province)
    avail_years = list(avail_years_tuple)

    if is_growth:
        periods_g = 4 if "yoy" in analysis_mode else 1
        df_g_all = _get_growth_df_cached(province, periods_g)

        if not df_g_all.empty:
            df_g_year = df_g_all[df_g_all["_tahun"] == int(year)]
            prefix_growth = prefix_hb if chosen_type == "HB" else prefix_hk

            if not df_g_year.empty:
                g_sector_col = next((c for c in df_g_year.columns if clean_sec in c.lower() and prefix_growth.lower() in c.lower()), None)
                g_p_col = next((c for c in df_g_year.columns if "penumpang" in c.lower()), None)
                g_bag_col = next((c for c in df_g_year.columns if "bagasi" in c.lower()), None)
                g_bar_col = next((c for c in df_g_year.columns if "barang" in c.lower()), None)

                growth_ts_data["pdrb_growth"] = df_g_year[g_sector_col].fillna(0).tolist() if g_sector_col and g_sector_col in df_g_year.columns else [0]*len(df_g_year)
                growth_ts_data["penumpang_growth"] = df_g_year[g_p_col].fillna(0).tolist() if g_p_col and g_p_col in df_g_year.columns else [0]*len(df_g_year)
                growth_ts_data["bagasi_growth"] = df_g_year[g_bag_col].fillna(0).tolist() if g_bag_col and g_bag_col in df_g_year.columns else [0]*len(df_g_year)
                growth_ts_data["barang_growth"] = df_g_year[g_bar_col].fillna(0).tolist() if g_bar_col and g_bar_col in df_g_year.columns else [0]*len(df_g_year)
            else:
                growth_mode_label = "YoY" if "yoy" in analysis_mode else "QoQ"
                data_warning = (
                    f"Tahun {year} tidak memiliki data pertumbuhan {growth_mode_label} "
                    f"(tahun awal dataset). Gunakan toggle 'Multi-Tahun (Pooled)' untuk regresi OLS."
                )

    triwulan_data = []
    for i, tw in enumerate(triwulan_labels):
        row_data: Dict[str, Any] = {
            "triwulan": tw,
            "pdrb_raw": float(pdrb_series.iloc[i]),
            "pdrb_idx": float(pdrb_idx[i]),
            "pdrb_qoq": float(pdrb_qoq[i]),
            "pdrb_hk_raw": float(pdrb_hk_series.iloc[i]),
            "pdrb_hk_idx": float(pdrb_hk_idx[i]),
            "pdrb_hk_qoq": float(pdrb_hk_qoq[i]),
            "pdrb_hb_raw": float(pdrb_hb_series.iloc[i]),
            "pdrb_hb_idx": float(pdrb_hb_idx[i]),
            "pdrb_hb_qoq": float(pdrb_hb_qoq[i]),
            "penumpang_raw": float(p_series.iloc[i]),
            "penumpang_idx": float(p_idx[i]),
            "penumpang_qoq": float(p_qoq[i]),
            "bagasi_raw": float(bag_series.iloc[i]),
            "bagasi_idx": float(bag_idx[i]),
            "bagasi_qoq": float(bag_qoq[i]),
            "barang_raw": float(bar_series.iloc[i]),
            "barang_idx": float(bar_idx[i]),
            "barang_qoq": float(bar_qoq[i]),
        }
        if growth_ts_data:
            row_data["pdrb_growth"] = float(growth_ts_data.get("pdrb_growth", [0]*4)[i]) if i < len(growth_ts_data.get("pdrb_growth", [])) else 0.0
            row_data["penumpang_growth"] = float(growth_ts_data.get("penumpang_growth", [0]*4)[i]) if i < len(growth_ts_data.get("penumpang_growth", [])) else 0.0
            row_data["bagasi_growth"] = float(growth_ts_data.get("bagasi_growth", [0]*4)[i]) if i < len(growth_ts_data.get("bagasi_growth", [])) else 0.0
            row_data["barang_growth"] = float(growth_ts_data.get("barang_growth", [0]*4)[i]) if i < len(growth_ts_data.get("barang_growth", [])) else 0.0

        triwulan_data.append(row_data)

    # 3. OLS Regression
    if transport_metric == "bagasi" and bag_col:
        y_raw_series = bag_series
        label_y = "Bagasi (Kg)"
    elif transport_metric == "barang" and bar_col:
        y_raw_series = bar_series
        label_y = "Barang/Kargo (Kg)"
    else:
        y_raw_series = p_series
        label_y = "Penumpang (Orang)"

    def compute_regression(x_s: pd.Series, y_s: pd.Series, label_x_str: str, label_y_str: str,
                           point_labels: Optional[List[str]] = None) -> Dict[str, Any]:
        x_vals = x_s.values.astype(float) if len(x_s) > 0 else np.array([], dtype=float)
        y_vals = y_s.values.astype(float) if len(y_s) > 0 else np.array([], dtype=float)
        valid_m = ~(np.isnan(x_vals) | np.isnan(y_vals) | np.isinf(x_vals) | np.isinf(y_vals))
        x_clean = x_vals[valid_m]
        y_clean = y_vals[valid_m]

        if point_labels is None:
            point_labels = triwulan_labels
        lbl_clean = [point_labels[k] for k in range(len(point_labels)) if k < len(valid_m) and valid_m[k]]

        res: Dict[str, Any] = {
            "slope": 0.0,
            "intercept": 0.0,
            "r_val": 0.0,
            "r_squared": 0.0,
            "p_value": 1.0,
            "equation": "y = 0x + 0",
            "label_x": label_x_str,
            "label_y": label_y_str,
            "points": [],
            "trend_line": [],
            "n_points": 0
        }

        if len(x_clean) >= 2:
            slope, intercept = np.polyfit(x_clean, y_clean, 1)
            r_matrix = np.corrcoef(x_clean, y_clean)
            r_val = float(r_matrix[0, 1]) if not np.isnan(r_matrix[0, 1]) else 0.0
            r_squared = float(r_val ** 2)

            try:
                _, p_val = pearsonr(x_clean, y_clean)
                p_val = float(p_val) if not np.isnan(p_val) else 1.0
            except Exception:
                p_val = 1.0

            sign_c = "+" if intercept >= 0 else "-"
            slope_str = f"{slope:.3e}" if abs(slope) < 0.0001 else f"{slope:.4f}"
            eq_str = f"y = {slope_str}x {sign_c} {abs(intercept):,.2f}"

            points = [{"triwulan": lbl_clean[k], "label": lbl_clean[k], "x": float(x_clean[k]), "y": float(y_clean[k])} for k in range(len(x_clean))]

            x_min, x_max = float(x_clean.min()), float(x_clean.max())
            trend_line = [
                {"x": x_min, "y": float(slope * x_min + intercept)},
                {"x": x_max, "y": float(slope * x_max + intercept)}
            ]

            res.update({
                "slope": float(slope),
                "intercept": float(intercept),
                "r_val": r_val,
                "r_squared": r_squared,
                "p_value": p_val,
                "equation": eq_str,
                "points": points,
                "trend_line": trend_line,
                "n_points": len(x_clean)
            })
        return res

    if is_growth:
        periods_g = 4 if "yoy" in analysis_mode else 1
        growth_mode_label = "YoY" if "yoy" in analysis_mode else "QoQ"
        pdrb_type_label = "HB" if chosen_type == "HB" else "HK"
        prefix_growth = prefix_hb if chosen_type == "HB" else prefix_hk

        label_y_growth = f"Pertumbuhan {label_y.split('(')[0].strip()} ({growth_mode_label} %)"
        label_x_growth = f"Pertumbuhan PDRB ({pdrb_type_label}): {active_label} ({growth_mode_label} %)"

        df_g_pooled = _get_growth_df_cached(province, periods_g)

        if not df_g_pooled.empty:
            g_sec_col = next((c for c in df_g_pooled.columns if clean_sec in c.lower() and prefix_growth.lower() in c.lower()), None)
            g_p_col = next((c for c in df_g_pooled.columns if "penumpang" in c.lower()), None)
            g_bag_col = next((c for c in df_g_pooled.columns if "bagasi" in c.lower()), None)
            g_bar_col = next((c for c in df_g_pooled.columns if "barang" in c.lower()), None)

            if transport_metric == "bagasi" and g_bag_col and g_bag_col in df_g_pooled.columns:
                g_y_col = g_bag_col
            elif transport_metric == "barang" and g_bar_col and g_bar_col in df_g_pooled.columns:
                g_y_col = g_bar_col
            elif g_p_col and g_p_col in df_g_pooled.columns:
                g_y_col = g_p_col
            else:
                g_y_col = None

            pooled_labels = [f"{int(r['_tahun'])} Q{int(r['_triwulan_num'])}" for _, r in df_g_pooled.iterrows()]
            if g_sec_col and g_sec_col in df_g_pooled.columns and g_y_col:
                reg_pooled = compute_regression(df_g_pooled[g_sec_col], df_g_pooled[g_y_col], label_x_growth, label_y_growth, pooled_labels)
            else:
                reg_pooled = compute_regression(pd.Series([], dtype=float), pd.Series([], dtype=float), label_x_growth, label_y_growth, [])

            df_g_year = df_g_pooled[df_g_pooled["_tahun"] == int(year)]
            year_labels = [f"{int(r['_tahun'])} Q{int(r['_triwulan_num'])}" for _, r in df_g_year.iterrows()]
            if not df_g_year.empty and g_sec_col and g_sec_col in df_g_year.columns and g_y_col:
                reg_year = compute_regression(df_g_year[g_sec_col], df_g_year[g_y_col], label_x_growth, label_y_growth, year_labels)
            else:
                reg_year = compute_regression(pd.Series([], dtype=float), pd.Series([], dtype=float), label_x_growth, label_y_growth, [])
        else:
            reg_pooled = compute_regression(pd.Series([], dtype=float), pd.Series([], dtype=float), label_x_growth, label_y_growth, [])
            reg_year = compute_regression(pd.Series([], dtype=float), pd.Series([], dtype=float), label_x_growth, label_y_growth, [])

    else:
        label_x_abs = f"PDRB: {active_label} ({actual_tipe})"
        label_y_abs = label_y

        year_labels = [f"{year} {tw}" for tw in triwulan_labels]
        reg_year = compute_regression(pdrb_series, y_raw_series, label_x_abs, label_y_abs, year_labels)

        # --- Hitung regresi terpisah untuk HK dan HB agar frontend tidak perlu fetch ulang ---
        label_x_hk = f"PDRB: {active_label} (HK)"
        label_x_hb = f"PDRB: {active_label} (HB)"
        if col_hk:
            reg_year_hk = compute_regression(pdrb_hk_series, y_raw_series, label_x_hk, label_y_abs, year_labels)
        else:
            reg_year_hk = reg_year
        if col_hb:
            reg_year_hb = compute_regression(pdrb_hb_series, y_raw_series, label_x_hb, label_y_abs, year_labels)
        else:
            reg_year_hb = reg_year

        if not df_multi_all.empty:
            my_p_col = next((c for c in df_multi_all.columns if "penumpang" in c.lower()), None)
            my_bag_col = next((c for c in df_multi_all.columns if "bagasi" in c.lower()), None)
            my_bar_col = next((c for c in df_multi_all.columns if "barang" in c.lower()), None)

            if transport_metric == "bagasi" and my_bag_col:
                my_y_col = my_bag_col
            elif transport_metric == "barang" and my_bar_col:
                my_y_col = my_bar_col
            elif my_p_col:
                my_y_col = my_p_col
            else:
                my_y_col = None

            pooled_labels = [f"{int(r['_tahun'])} Q{int(r['_triwulan_num'])}" for _, r in df_multi_all.iterrows()]

            # Pooled regression untuk tipe aktif
            prefix_active = prefix_hb if actual_tipe == "HB" else prefix_hk
            my_sec_col = next((c for c in df_multi_all.columns if clean_sec in c.lower() and prefix_active.lower() in c.lower()), None)
            if my_sec_col and my_sec_col in df_multi_all.columns and my_y_col:
                reg_pooled = compute_regression(df_multi_all[my_sec_col], df_multi_all[my_y_col], f"{label_x_abs} [Pooled]", label_y_abs, pooled_labels)
            else:
                reg_pooled = compute_regression(pd.Series([], dtype=float), pd.Series([], dtype=float), label_x_abs, label_y_abs, [])

            # Pooled regression terpisah HK & HB
            my_sec_hk = next((c for c in df_multi_all.columns if clean_sec in c.lower() and prefix_hk.lower() in c.lower()), None)
            my_sec_hb = next((c for c in df_multi_all.columns if clean_sec in c.lower() and prefix_hb.lower() in c.lower()), None)
            if my_sec_hk and my_sec_hk in df_multi_all.columns and my_y_col:
                reg_pooled_hk = compute_regression(df_multi_all[my_sec_hk], df_multi_all[my_y_col], f"{label_x_hk} [Pooled]", label_y_abs, pooled_labels)
            else:
                reg_pooled_hk = reg_pooled
            if my_sec_hb and my_sec_hb in df_multi_all.columns and my_y_col:
                reg_pooled_hb = compute_regression(df_multi_all[my_sec_hb], df_multi_all[my_y_col], f"{label_x_hb} [Pooled]", label_y_abs, pooled_labels)
            else:
                reg_pooled_hb = reg_pooled
        else:
            reg_pooled = reg_year
            reg_pooled_hk = reg_year_hk
            reg_pooled_hb = reg_year_hb

    reg_result = reg_pooled if str(ols_scope).lower() == "pooled" else reg_year

    # Untuk mode growth, HK/HB sudah ditentukan oleh analysis_mode — gunakan reg_result
    if is_growth:
        reg_hk = reg_result
        reg_hb = reg_result
        reg_annual = reg_pooled
        reg_annual_hk = reg_pooled
        reg_annual_hb = reg_pooled
    else:
        # Untuk mode absolut, kirim regresi HK & HB terpisah
        reg_hk = reg_year_hk
        reg_hb = reg_year_hb
        reg_annual = reg_pooled
        reg_annual_hk = reg_pooled_hk
        reg_annual_hb = reg_pooled_hb

    # 4. Sektor Lapangan Usaha Top 10 Summary
    target_lu_cols = lu_hb_cols if actual_tipe == "HB" else lu_hk_cols
    sectors_summary = []
    for col in target_lu_cols:
        clean_name = re.sub(r"^LU\s*\([A-Z]+\)\s*-\s*", "", col)
        if "produk domestik" not in clean_name.lower() and "pdrb" not in clean_name.lower():
            tot_val = float(df_tri[col].sum())
            sectors_summary.append({"sector": clean_name, "total_output": tot_val})

    sectors_summary = sorted(sectors_summary, key=lambda x: x["total_output"], reverse=True)
    total_lu_sum = sum(s["total_output"] for s in sectors_summary) or 1.0
    for s in sectors_summary:
        s["percentage"] = round((s["total_output"] / total_lu_sum) * 100.0, 2)

    response_data: Dict[str, Any] = {
        "province": province,
        "year": year,
        "active_sector": active_label,
        "category": category,
        "tipe_pdrb": actual_tipe,
        "has_hk": col_hk is not None,
        "has_hb": col_hb is not None,
        "analysis_mode": analysis_mode,
        "ols_scope": ols_scope,
        "kpi": kpi,
        "triwulan_data": triwulan_data,
        "regression": reg_result,
        "regression_year": reg_year,
        "regression_pooled": reg_pooled,
        "regression_hk": reg_hk,
        "regression_hb": reg_hb,
        "regression_annual": reg_annual,
        "regression_annual_hk": reg_annual_hk,
        "regression_annual_hb": reg_annual_hb,
        "sectors_summary": sectors_summary[:10],
        "all_sectors_summary": sectors_summary,
        "available_years": avail_years,
    }
    if data_warning:
        response_data["data_warning"] = data_warning
    if is_growth and growth_ts_data:
        response_data["growth_ts_available"] = True

    return response_data


@app.get("/api/data")
async def get_dashboard_data(
    province: str = "sulawesi_selatan",
    year: str = "2024",
    sector: str = "Transportasi dan Pergudangan",
    category: str = "lu",
    tipe_pdrb: str = "HK",
    transport_metric: str = "penumpang",
    analysis_mode: str = "growth_yoy",
    ols_scope: str = "year"
):
    """Endpoint agregasi lengkap dengan in-memory LRU caching berkecepatan tinggi (<10ms)."""
    res = _compute_dashboard_data(
        province, year, sector, category, tipe_pdrb, transport_metric, analysis_mode, ols_scope
    )
    if "error" in res:
        raise HTTPException(status_code=404, detail=res["error"])
    return cached_json_response(res)


@functools.lru_cache(maxsize=64)
def _compute_raw_sheet(province: str, year: str, sheet: str) -> Dict[str, Any]:
    stem = f"{province}_{year}_{sheet}"
    df = get_parquet_df(stem)
    if df is None:
        return {"error": f"Sheet {stem}.parquet tidak ditemukan di {DATA_DIR}."}
    
    df_clean = df.replace({np.nan: None})
    return {
        "stem": stem,
        "columns": df_clean.columns.tolist(),
        "rows": df_clean.to_dict(orient="records"),
        "count": len(df_clean)
    }


@app.get("/api/raw_sheet")
async def get_raw_sheet(
    province: str = "sulawesi_selatan",
    year: str = "2024",
    sheet: str = "pdrb_triwulan"
):
    """Membaca raw sheet parquet tabular untuk ekspor dan inspeksi dengan caching."""
    res = _compute_raw_sheet(province, year, sheet)
    if "error" in res:
        raise HTTPException(status_code=404, detail=res["error"])
    return cached_json_response(res)


# ======================================================================================
# MULTICOLLINEARITY & VIF DIAGNOSTIC ENGINE (Pure NumPy Matrix Algebra)
# ======================================================================================

def _normalize_name(s: str) -> str:
    """Membersihkan dan menormalisasi string nama sektor untuk pencocokan toleran tanda baca."""
    s = re.sub(r'^(LU|Peng)\s*\([A-Z]+\)\s*-\s*', '', s, flags=re.IGNORECASE)
    s = re.sub(r'[^\w\s]', ' ', s).lower()
    return ' '.join(s.split())


_SECTOR_ALIASES: Dict[str, str] = {
    'ekspor luar negeri': 'ekspor barang dan jasa',
    'impor luar negeri': 'impor barang dan jasa',
}


class MulticoRequest(BaseModel):
    """Schema input untuk endpoint POST multikolinearitas."""
    province: str = "sulawesi_selatan"
    price_type: str = "HK"
    analysis_mode: str = "growth_yoy"
    category: str = "lu"
    sectors: Optional[List[str]] = None


def _find_optimal_vif_subset(
    df_data: pd.DataFrame,
    matched_cols: List[str],
    matched_orig_sectors: List[str],
    max_k: int
) -> List[str]:
    """
    Secara iteratif mengeliminasi variabel dengan VIF tertinggi
    hingga menyisakan subset sektor dengan seluruh VIF < 5 (maksimal k = max_k sektor).
    """
    cur_cols = list(matched_cols)
    cur_orig = list(matched_orig_sectors)

    if len(cur_cols) <= 2:
        return cur_orig

    while len(cur_cols) > 2:
        sub = df_data[cur_cols].dropna()
        if len(sub) < 3:
            break

        R_cur = sub.corr().values
        R_cur = np.nan_to_num(R_cur, nan=0.0)
        np.fill_diagonal(R_cur, 1.0)
        k_cur = len(cur_cols)
        N_cur = len(sub)

        try:
            cond_cur = float(np.linalg.cond(R_cur))
        except Exception:
            cond_cur = 1.0

        use_ridge = (k_cur >= N_cur - 1) or (cond_cur > 1e3)
        if use_ridge:
            lam = 0.05
            R_r = R_cur + lam * np.eye(k_cur)
            inv_r = np.linalg.inv(R_r)
            cur_vifs = np.diag(inv_r @ R_cur @ inv_r)
        else:
            try:
                cur_vifs = np.diag(np.linalg.inv(R_cur))
            except Exception:
                lam = 0.05
                R_r = R_cur + lam * np.eye(k_cur)
                inv_r = np.linalg.inv(R_r)
                cur_vifs = np.diag(inv_r @ R_cur @ inv_r)

        max_idx = int(np.argmax(cur_vifs))
        max_val = float(cur_vifs[max_idx])

        # Kondisi berhenti: jika k <= max_k dan seluruh VIF < 5.0 dan cond <= 30.0
        if k_cur <= max_k and max_val < 5.0 and cond_cur <= 30.0:
            break

        # Jika masih ada VIF >= 5.0 atau k > max_k, eliminasi sektor dengan VIF tertinggi
        del cur_cols[max_idx]
        del cur_orig[max_idx]

    return cur_orig


@functools.lru_cache(maxsize=256)
def _compute_multicollinearity_cached(
    province: str,
    price_type: str,
    analysis_mode: str,
    category: str,
    sectors_tuple: Tuple[str, ...]
) -> Dict[str, Any]:
    """
    Kalkulasi Diagnostik Multikolinearitas Antar-Sektor:
    1. Matriks Korelasi Pearson R (k x k)
    2. Generalized Ridge VIF (saat k >= N-1 atau matrix singular) vs Standard Inversion
    3. Variance Inflation Factor (VIF_j) & Tolerance (1 / VIF_j)
    4. Condition Number (np.linalg.cond(R))
    5. Auto-Pruning Subset Bebas Multiko (VIF < 5)
    6. Rekomendasi Ekonometrika Otomatis
    """
    sectors = list(sectors_tuple)
    if not sectors or len(sectors) < 2:
        return {
            "status": "error",
            "message": "Minimal 2 sektor harus dipilih untuk kalkulasi multikolinearitas (VIF & Korelasi Antar-Sektor)."
        }

    is_growth = analysis_mode.startswith("growth_")
    periods = 4 if "yoy" in analysis_mode else 1
    pt = "HB" if (price_type.upper() == "HB" or analysis_mode.endswith("_hb") or analysis_mode == "abs_hb") else "HK"

    if is_growth:
        df_data = _get_growth_df_cached(province, periods)
    else:
        df_data, _ = _build_multi_year_df_cached(province)

    if df_data.empty:
        return {
            "status": "error",
            "message": f"Data multi-tahun untuk provinsi {province} tidak ditemukan."
        }

    prefix = ("LU (HB)" if pt == "HB" else "LU (HK)") if category == "lu" else ("Peng (HB)" if pt == "HB" else "Peng (HK)")
    available_cols = [c for c in df_data.columns if c.lower().startswith(prefix.lower())]

    # Pencocokan nama sektor dengan kolom aktual pada DataFrame secara toleran tanda baca (koma, titik koma, dsb.)
    matched_cols = []
    matched_labels = []
    matched_orig_sectors = []

    for sec in sectors:
        norm_sec = _normalize_name(sec)
        alias_sec = _SECTOR_ALIASES.get(norm_sec, norm_sec)

        found_col = None
        # 1. Exact match setelah normalisasi
        for c in available_cols:
            norm_c = _normalize_name(c)
            if norm_sec == norm_c or alias_sec == norm_c:
                found_col = c
                break

        # 2. Contains match fallback
        if not found_col:
            for c in available_cols:
                norm_c = _normalize_name(c)
                if norm_sec in norm_c or norm_c in norm_sec or alias_sec in norm_c or norm_c in alias_sec:
                    found_col = c
                    break

        if found_col and found_col not in matched_cols:
            matched_cols.append(found_col)
            # Format label bersih
            clean_label = re.sub(r'^(LU|Peng)\s*\([A-Z]+\)\s*-\s*', '', found_col).strip()
            matched_labels.append(clean_label)
            matched_orig_sectors.append(sec)

    if len(matched_cols) < 2:
        return {
            "status": "error",
            "message": f"Hanya ditemukan {len(matched_cols)} kolom valid dari {len(sectors)} sektor yang dipilih. Minimal dibutuhkan 2 sektor yang cocok."
        }

    sub_df = df_data[matched_cols].dropna()
    if len(sub_df) < 3:
        return {
            "status": "error",
            "message": f"Jumlah observasi valid ({len(sub_df)} baris) tidak mencukupi untuk estimasi matriks korelasi."
        }

    # 1. Matriks Korelasi Pearson R (k x k)
    R = sub_df.corr().values
    R = np.nan_to_num(R, nan=0.0)
    np.fill_diagonal(R, 1.0)

    k = len(matched_labels)
    N = len(sub_df)

    # 4. Condition Number
    try:
        cond_num = float(np.linalg.cond(R))
    except Exception:
        cond_num = 1.0

    # 2. Invers Matriks Korelasi & Generalized Ridge VIF
    # Saat k >= N - 1 atau Condition Number > 1e3, gunakan formula Generalized Ridge VIF
    is_ridge = (k >= N - 1) or (cond_num > 1e3)
    lam = 0.05

    if is_ridge:
        try:
            R_ridge = R + lam * np.eye(k)
            inv_ridge = np.linalg.inv(R_ridge)
            vifs = np.diag(inv_ridge @ R @ inv_ridge)
            method_str = "Ridge Regularized VIF (λ=0.05)"
        except Exception:
            inv_ridge = np.linalg.pinv(R + lam * np.eye(k))
            vifs = np.diag(inv_ridge @ R @ inv_ridge)
            method_str = "Ridge Regularized VIF (λ=0.05)"
    else:
        try:
            R_inv = np.linalg.inv(R)
            vifs = np.diag(R_inv)
            method_str = "Standard Inversion OLS"
        except np.linalg.LinAlgError:
            R_ridge = R + lam * np.eye(k)
            inv_ridge = np.linalg.inv(R_ridge)
            vifs = np.diag(inv_ridge @ R @ inv_ridge)
            method_str = "Ridge Regularized VIF (λ=0.05)"
        except Exception as e:
            return {
                "status": "error",
                "message": f"Gagal menghitung invers matriks korelasi: {e}"
            }

    # Hitung Auto-Pruning Subset Bebas Multikolinearitas (VIF < 5, max k = N - 2)
    max_rec_k = max(2, min(N - 2, k))
    optimal_subset = _find_optimal_vif_subset(df_data, matched_cols, matched_orig_sectors, max_rec_k)

    vif_results = []
    has_multicollinearity = is_ridge or (cond_num > 30.0)
    high_corr_pairs = []

    for i in range(k):
        vif_val = float(vifs[i])
        # Koreksi floating point
        if vif_val < 1.0:
            vif_val = 1.0
        tol_val = float(1.0 / vif_val) if vif_val > 0 else 0.0

        if vif_val >= 10.0:
            status_str = "Multiko Berat (≥ 10)"
            badge_color = "red"
            has_multicollinearity = True
        elif vif_val >= 5.0:
            status_str = "Multiko Sedang (5 - 10)"
            badge_color = "yellow"
            has_multicollinearity = True
        else:
            if is_ridge:
                status_str = "Ridge Terkoreksi (< 5)"
                badge_color = "green"
            else:
                status_str = "Aman (< 5)"
                badge_color = "green"

        vif_results.append({
            "sector": matched_labels[i],
            "vif": round(vif_val, 3),
            "tolerance": round(tol_val, 4),
            "status": status_str,
            "badge_color": badge_color
        })

    # Evaluasi Condition Number
    if cond_num > 30.0:
        has_multicollinearity = True
    elif cond_num > 15.0 and any(v["vif"] >= 5.0 for v in vif_results):
        has_multicollinearity = True

    # 5. Deteksi Pasangan Sektor Berkorelasi Sangat Tinggi (|r| >= 0.80)
    for i in range(k):
        for j in range(i + 1, k):
            r_val = float(R[i, j])
            if abs(r_val) >= 0.80:
                high_corr_pairs.append({
                    "sector_a": matched_labels[i],
                    "sector_b": matched_labels[j],
                    "r": round(r_val, 4)
                })

    # 6. Formulasi Rekomendasi Ekonometrika
    recommendations = []
    if is_ridge:
        recommendations.append(
            f"Jumlah prediktor (k={k}) mendekati/melebihi derajat bebas data (N={N}, Condition Number = {cond_num:.2e}). "
            f"Perhitungan VIF distabilkan menggunakan formula Generalized Ridge Regularization (λ=0.05)."
        )
    elif cond_num > 30.0:
        recommendations.append(
            f"Condition Number bernilai {cond_num:.2f} (> 30), mengindikasikan struktur data mengalami multikolinearitas berat."
        )
    elif cond_num > 15.0:
        recommendations.append(
            f"Condition Number bernilai {cond_num:.2f} (15–30), mengindikasikan gejala multikolinearitas moderat."
        )

    if optimal_subset and len(optimal_subset) < k:
        recommendations.append(
            f"Subset rekomendasi bebas multiko ({len(optimal_subset)} sektor, VIF < 5): {', '.join(optimal_subset)}."
        )

    if high_corr_pairs:
        # Urutkan berdasarkan nilai korelasi absolut tertinggi
        high_corr_pairs.sort(key=lambda x: abs(x["r"]), reverse=True)
        top_pair = high_corr_pairs[0]
        recommendations.append(
            f"Sektor '{top_pair['sector_a']}' dan '{top_pair['sector_b']}' memiliki korelasi sangat tinggi (r = {top_pair['r']:.2f}). "
            f"Disarankan hanya memasukkan salah satu sektor ke dalam model regresi berganda."
        )
        if len(high_corr_pairs) > 1:
            other_pairs_str = ", ".join([f"'{p['sector_a']}' & '{p['sector_b']}' (r={p['r']:.2f})" for p in high_corr_pairs[1:3]])
            recommendations.append(f"Pasangan lain dengan korelasi tinggi: {other_pairs_str}.")

    # Rekomendasi sektor VIF tertinggi jika ada multiko berat
    severe_vifs = [v for v in vif_results if v["vif"] >= 10.0]
    if severe_vifs:
        max_vif_sec = max(severe_vifs, key=lambda x: x["vif"])
        recommendations.append(
            f"Sektor '{max_vif_sec['sector']}' memiliki VIF tertinggi ({max_vif_sec['vif']:.2f}). "
            f"Pertimbangkan untuk mengeluarkan variabel ini atau menggunakan metode regularisasi (Ridge/Lasso Regression)."
        )

    if not has_multicollinearity and not high_corr_pairs and not is_ridge:
        recommendation_str = (
            "Tidak ditemukan gejala multikolinearitas yang signifikan (seluruh VIF < 5 dan Condition Number aman). "
            "Seluruh sektor yang dipilih dapat diikutsertakan bersamaan dalam model regresi berganda."
        )
    else:
        recommendation_str = " ".join(recommendations)

    # 7. Kalkulasi Multikolinearitas Antar-Indikator Transportasi (Penumpang, Bagasi, Barang)
    p_col = next((c for c in df_data.columns if "penumpang" in c.lower()), None)
    bag_col = next((c for c in df_data.columns if "bagasi" in c.lower()), None)
    bar_col = next((c for c in df_data.columns if "barang" in c.lower()), None)

    valid_trans = [("Penumpang", p_col), ("Bagasi", bag_col), ("Barang", bar_col)]
    trans_labels = [l for l, c in valid_trans if c is not None and c in df_data.columns]
    trans_cols = [c for l, c in valid_trans if c is not None and c in df_data.columns]

    transport_multicollinearity = {
        "condition_number": 1.0,
        "has_multicollinearity": False,
        "vif_results": [],
        "matrix": {
            "labels": trans_labels,
            "matrix": []
        },
        "recommendation": "Data indikator transportasi tidak mencukupi untuk evaluasi multikolinearitas."
    }

    if len(trans_cols) >= 2:
        sub_t = df_data[trans_cols].dropna()
        if len(sub_t) >= 3:
            R_t = sub_t.corr().values
            R_t = np.nan_to_num(R_t, nan=0.0)
            np.fill_diagonal(R_t, 1.0)
            try:
                R_t_inv = np.linalg.pinv(R_t)
                vifs_t = np.diag(R_t_inv)
            except Exception:
                vifs_t = [1.0] * len(trans_cols)

            try:
                cond_num_t = float(np.linalg.cond(R_t))
            except Exception:
                cond_num_t = 1.0

            trans_vif_results = []
            has_multi_t = False
            for idx_t, t_label in enumerate(trans_labels):
                val_t = float(vifs_t[idx_t])
                if val_t < 1.0:
                    val_t = 1.0
                tol_t = float(1.0 / val_t) if val_t > 0 else 0.0

                if val_t >= 10.0:
                    status_t = "Multiko Berat (≥ 10)"
                    badge_t = "red"
                    has_multi_t = True
                elif val_t >= 5.0:
                    status_t = "Multiko Sedang (5 - 10)"
                    badge_t = "yellow"
                    has_multi_t = True
                else:
                    status_t = "Aman (< 5)"
                    badge_t = "green"

                trans_vif_results.append({
                    "indicator": t_label,
                    "vif": round(val_t, 3),
                    "tolerance": round(tol_t, 4),
                    "status": status_t,
                    "badge_color": badge_t
                })

            if cond_num_t > 30.0 or (cond_num_t > 15.0 and any(v["vif"] >= 5.0 for v in trans_vif_results)):
                has_multi_t = True

            # Formulasi rekomendasi transportasi
            rec_t = []
            if has_multi_t:
                rec_t.append(
                    f"Terdeteksi multikolinearitas antar-indikator transportasi (Condition Number = {cond_num_t:.2f})."
                )
                r_pb = float(R_t[0, 1]) if len(trans_labels) >= 2 else 0.0
                if abs(r_pb) >= 0.8:
                    rec_t.append(
                        f"Indikator '{trans_labels[0]}' dan '{trans_labels[1]}' memiliki korelasi linear sangat kuat (r = {r_pb:.2f}). "
                        f"Hindari memasukkan keduanya secara bersamaan sebagai prediktor simultan dalam satu model regresi OLS; "
                        f"disarankan memilih salah satu target atau menerapkan reduksi dimensi (PCA)."
                    )
            else:
                rec_t.append(
                    "Indikator transportasi (Penumpang, Bagasi, Barang) memiliki VIF < 5 dan Condition Number aman. "
                    "Seluruh indikator aman digabungkan bersamaan dalam analisis regresi multivariat."
                )

            transport_multicollinearity = {
                "condition_number": round(cond_num_t, 2),
                "has_multicollinearity": has_multi_t,
                "vif_results": trans_vif_results,
                "matrix": {
                    "labels": trans_labels,
                    "matrix": [[round(float(val), 4) for val in row] for row in R_t]
                },
                "recommendation": " ".join(rec_t)
            }

    # 8. Matriks Korelasi Silang (PDRB Sektor × Indikator Transportasi)
    cross_matrix = []
    cross_p_values = []
    strongest_targets = []

    for idx_s, sec_col in enumerate(matched_cols):
        sec_r_row = []
        sec_p_row = []
        best_t = None
        best_r = None
        best_p = 1.0

        for t_col, t_label in zip(trans_cols, trans_labels):
            sub_cross = df_data[[sec_col, t_col]].dropna()
            if len(sub_cross) >= 3:
                r_cross, p_cross = pearsonr(sub_cross[sec_col], sub_cross[t_col])
                if np.isnan(r_cross):
                    r_cross = 0.0
                    p_cross = 1.0
            else:
                r_cross, p_cross = 0.0, 1.0

            r_val_rounded = round(float(r_cross), 4)
            p_val_rounded = round(float(p_cross), 4)
            sec_r_row.append(r_val_rounded)
            sec_p_row.append(p_val_rounded)

            if best_r is None or abs(r_cross) > abs(best_r):
                best_r = r_cross
                best_p = p_cross
                best_t = t_label

        cross_matrix.append(sec_r_row)
        cross_p_values.append(sec_p_row)
        strongest_targets.append({
            "sector": matched_labels[idx_s],
            "best_target": best_t or "Penumpang",
            "r": round(float(best_r if best_r is not None else 0.0), 4),
            "p": round(float(best_p), 4)
        })

    cross_transport_matrix = {
        "sector_labels": matched_labels,
        "transport_labels": trans_labels,
        "matrix": cross_matrix,
        "p_values": cross_p_values,
        "strongest_targets": strongest_targets
    }

    return {
        "status": "success",
        "method": method_str,
        "condition_number": round(cond_num, 2),
        "has_multicollinearity": has_multicollinearity,
        "is_rank_deficient": is_ridge,
        "optimal_subset": optimal_subset,
        "vif_results": vif_results,
        "inter_sector_matrix": {
            "labels": matched_labels,
            "matrix": [[round(float(val), 4) for val in row] for row in R]
        },
        "cross_transport_matrix": cross_transport_matrix,
        "transport_multicollinearity": transport_multicollinearity,
        "recommendation": recommendation_str,
        "recommendations_list": recommendations
    }


def compute_multicollinearity(
    province: str = "sulawesi_selatan",
    price_type: str = "HK",
    analysis_mode: str = "growth_yoy",
    category: str = "lu",
    sectors: Optional[List[str]] = None
) -> Dict[str, Any]:
    """Public wrapper untuk kalkulasi multikolinearitas."""
    if not sectors:
        # Default core sectors
        if category == "lu":
            sectors = [
                "Industri Pengolahan",
                "Konstruksi",
                "Perdagangan Besar dan Eceran, Reparasi Mobil dan Sepeda Motor",
                "Transportasi dan Pergudangan",
                "Penyediaan Akomodasi dan Makan Minum",
                "Pengadaan Air, Pengelolaan Sampah, Limbah, dan Daur Ulang"
            ]
        else:
            sectors = [
                "Pengeluaran Konsumsi Rumah Tangga",
                "Pembentukan Modal Tetap Bruto",
                "Ekspor Barang dan Jasa",
                "Impor Barang dan Jasa"
            ]

    # Filter empty / whitespace-only entries
    clean_sectors = [s.strip() for s in sectors if s and s.strip()]

    sectors_tuple = tuple(sorted(clean_sectors))
    return _compute_multicollinearity_cached(
        province=province,
        price_type=price_type,
        analysis_mode=analysis_mode,
        category=category,
        sectors_tuple=sectors_tuple
    )


@app.post("/api/multicollinearity")
async def post_multicollinearity(req: MulticoRequest):
    """
    Endpoint POST Diagnostik Multikolinearitas (VIF & Korelasi Antar-Sektor).
    Menerima body JSON dengan schema MulticoRequest agar nama sektor dengan tanda koma
    tidak terpecah saat parsing.
    """
    res = compute_multicollinearity(
        province=req.province,
        price_type=req.price_type,
        analysis_mode=req.analysis_mode,
        category=req.category,
        sectors=req.sectors
    )
    if res.get("status") == "error":
        return JSONResponse(status_code=400, content=res, headers=CACHE_HEADERS)
    return cached_json_response(res)


@app.get("/api/multicollinearity")
async def get_multicollinearity(
    province: str = "sulawesi_selatan",
    price_type: str = "HK",
    analysis_mode: str = "growth_yoy",
    category: str = "lu",
    sectors: Optional[List[str]] = Query(default=None)
):
    """
    Endpoint GET Diagnostik Multikolinearitas (VIF & Korelasi Antar-Sektor).
    Mendukung query parameter untuk backward-compatibility.
    """
    res = compute_multicollinearity(
        province=province,
        price_type=price_type,
        analysis_mode=analysis_mode,
        category=category,
        sectors=sectors
    )
    if res.get("status") == "error":
        return JSONResponse(status_code=400, content=res, headers=CACHE_HEADERS)
    return cached_json_response(res)


# ======================================================================================
# MULTI-SHEET EXCEL REPORT GENERATOR (.xlsx with openpyxl)
# ======================================================================================

class ExportExcelRequest(BaseModel):
    """Schema input untuk endpoint POST ekspor laporan Excel multi-sheet."""
    province: str = "sulawesi_selatan"
    compare_province: str = "gorontalo"
    year: str = "2024"
    analysis_mode: str = "growth_yoy"
    transport_type: str = "penumpang"


def generate_full_excel_report(
    province: str = "sulawesi_selatan",
    compare_province: str = "gorontalo",
    year: str = "2024",
    analysis_mode: str = "growth_yoy",
    transport_type: str = "penumpang"
) -> io.BytesIO:
    """
    Menghasilkan Workbook Excel multi-sheet (.xlsx) terintegrasi dengan 5 sheet analisis:
    1. 1_Korelasi_PDRB (17 Sektor LU & 7 Pengeluaran vs Transportasi + R²)
    2. 2_Diagnostik_VIF (VIF 6 Sektor Inti, Matriks Rx, & Multikolinearitas Transportasi)
    3. 3_Komparasi_Wilayah (Head-to-Head Provinsi A vs B: r, R², |Δr|, VIF)
    4. 4_Tren_Triwulanan (Time-Series 2020-2024: Level, YoY %, QoQ %)
    5. 5_Data_Mentah (Tabel Nilai Riil 17 Sektor ADHK & ADHB)
    """
    manifest = load_manifest()
    prov_dict = manifest.get("provinces", {})
    prov_a_name = prov_dict.get(province, {}).get("name", province.replace("_", " ").title())
    prov_b_name = prov_dict.get(compare_province, {}).get("name", compare_province.replace("_", " ").title())

    mode_labels = {
        "growth_yoy_all": "Pertumbuhan YoY (%) — HK & HB (Semua Tipe)",
        "growth_yoy": "Pertumbuhan YoY (%) — Harga Konstan (Riil)",
        "growth_yoy_hb": "Pertumbuhan YoY (%) — Harga Berlaku (Nominal)",
        "growth_qoq_all": "Pertumbuhan QoQ (%) — HK & HB (Semua Tipe)",
        "growth_qoq": "Pertumbuhan QoQ (%) — Harga Konstan (Riil)",
        "growth_qoq_hb": "Pertumbuhan QoQ (%) — Harga Berlaku (Nominal)",
        "abs_all": "Nilai Level / Absolut — Semua Tipe (HK & HB)",
        "abs_hk": "Nilai Level / Absolut — Harga Konstan (Riil)",
        "abs_hb": "Nilai Level / Absolut — Harga Berlaku (Nominal)"
    }
    analysis_label = mode_labels.get(analysis_mode, analysis_mode)

    trans_labels = {
        "penumpang": "✈️ Penumpang (Orang)",
        "bagasi": "🧳 Bagasi (Kg)",
        "barang": "📦 Barang / Kargo (Kg)"
    }
    trans_label = trans_labels.get(transport_type.lower(), transport_type.title())

    # --- STYLING DEFINITIONS ---
    navy_header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    navy_title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    sky_section_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    indigo_section_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    amber_section_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
    emerald_section_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    zebra_light_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    title_font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    subtitle_font = Font(name="Calibri", size=9.5, italic=True, color="CBD5E1")
    section_font = Font(name="Calibri", size=10.5, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=9.5, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=9.5, color="0F172A")
    data_bold_font = Font(name="Calibri", size=9.5, bold=True, color="0F172A")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    header_border = Border(
        left=Side(style='thin', color='475569'),
        right=Side(style='thin', color='475569'),
        top=Side(style='medium', color='0F172A'),
        bottom=Side(style='medium', color='0F172A')
    )

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def autofit_columns(ws, max_cols=None):
        cols_to_check = ws.columns if max_cols is None else list(ws.columns)[:max_cols]
        for col in cols_to_check:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row in [1, 2, 3, 4] and cell.coordinate in ws.merged_cells:
                    continue
                v = cell.value
                if v is not None:
                    s = str(v)
                    if '\n' in s:
                        s = max(s.split('\n'), key=len)
                    if len(s) > max_len:
                        max_len = len(s)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    # =========================================================================
    # SHEET 1: 1_Korelasi_PDRB
    # =========================================================================
    ws1 = wb.create_sheet(title="1_Korelasi_PDRB")
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells("A1:L1")
    ws1["A1"] = "LAPORAN ANALISIS KORELASI PDRB DAN TRANSPORTASI UDARA"
    ws1["A1"].font = title_font
    ws1["A1"].fill = navy_title_fill
    ws1["A1"].alignment = align_center
    ws1.row_dimensions[1].height = 28

    ws1.merge_cells("A2:L2")
    ws1["A2"] = f"Provinsi: {prov_a_name} | Tahun: {year} | Mode: {analysis_label} | Sumber Data: BPS RI"
    ws1["A2"].font = subtitle_font
    ws1["A2"].fill = navy_header_fill
    ws1["A2"].alignment = align_center
    ws1.row_dimensions[2].height = 18

    curr_row = 4

    # Bagian A: 17 Sektor Lapangan Usaha
    ws1.merge_cells(f"A{curr_row}:L{curr_row}")
    ws1[f"A{curr_row}"] = "BAGIAN A: 17 SEKTOR PDRB MENURUT LAPANGAN USAHA (LU) VS TRANSPORTASI"
    ws1[f"A{curr_row}"].font = section_font
    ws1[f"A{curr_row}"].fill = sky_section_fill
    ws1[f"A{curr_row}"].alignment = align_left
    ws1.row_dimensions[curr_row].height = 22
    curr_row += 1

    headers_correl = [
        "No", "Sektor PDRB", "Tipe PDRB",
        "r Penumpang", "R² Penumpang", "p-value Penumpang",
        "r Bagasi", "R² Bagasi", "p-value Bagasi",
        "r Barang", "R² Barang", "p-value Barang"
    ]
    for c_idx, h in enumerate(headers_correl, 1):
        cell = ws1.cell(row=curr_row, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align_center
        cell.border = header_border
    ws1.row_dimensions[curr_row].height = 24
    curr_row += 1

    corr_lu = _compute_correlations(province, year, type="lu", filter_mode="all", sort_by="default", pdrb_type="all", search="", analysis_mode=analysis_mode)
    lu_rows = corr_lu.get("rows", [])

    for idx, r in enumerate(lu_rows, 1):
        sec_name = r.get("Lapangan Usaha", "-")
        pt = r.get("Tipe PDRB", "HK")
        rp = r.get("Korelasi dgn Penumpang")
        r2p = (rp ** 2) if (rp is not None and not np.isnan(rp)) else None
        pp = r.get("p-value Penumpang")

        rbag = r.get("Korelasi dgn Bagasi")
        r2bag = (rbag ** 2) if (rbag is not None and not np.isnan(rbag)) else None
        pbag = r.get("p-value Bagasi")

        rbar = r.get("Korelasi dgn Barang")
        r2bar = (rbar ** 2) if (rbar is not None and not np.isnan(rbar)) else None
        pbar = r.get("p-value Barang")

        fill_row = zebra_light_fill if idx % 2 == 0 else white_fill
        row_vals = [idx, sec_name, pt, rp, r2p, pp, rbag, r2bag, pbag, rbar, r2bar, pbar]

        for c_idx, v in enumerate(row_vals, 1):
            cell = ws1.cell(row=curr_row, column=c_idx, value=v)
            cell.font = data_font
            cell.fill = fill_row
            cell.border = thin_border
            if c_idx == 1:
                cell.alignment = align_center
            elif c_idx in [2, 3]:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                if c_idx in [4, 7, 10]:
                    cell.number_format = "0.0000"
                elif c_idx in [5, 8, 11]:
                    cell.number_format = "0.0%"
                elif c_idx in [6, 9, 12]:
                    cell.number_format = "0.0000"
        ws1.row_dimensions[curr_row].height = 18
        curr_row += 1

    curr_row += 1
    # Bagian B: 7 Komponen Pengeluaran
    ws1.merge_cells(f"A{curr_row}:L{curr_row}")
    ws1[f"A{curr_row}"] = "BAGIAN B: 7 KOMPONEN PDRB MENURUT PENGELUARAN VS TRANSPORTASI"
    ws1[f"A{curr_row}"].font = section_font
    ws1[f"A{curr_row}"].fill = indigo_section_fill
    ws1[f"A{curr_row}"].alignment = align_left
    ws1.row_dimensions[curr_row].height = 22
    curr_row += 1

    headers_peng = [
        "No", "Komponen Pengeluaran", "Tipe PDRB",
        "r Penumpang", "R² Penumpang", "p-value Penumpang",
        "r Bagasi", "R² Bagasi", "p-value Bagasi",
        "r Barang", "R² Barang", "p-value Barang"
    ]
    for c_idx, h in enumerate(headers_peng, 1):
        cell = ws1.cell(row=curr_row, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align_center
        cell.border = header_border
    ws1.row_dimensions[curr_row].height = 24
    curr_row += 1

    corr_peng = _compute_correlations(province, year, type="peng", filter_mode="all", sort_by="default", pdrb_type="all", search="", analysis_mode=analysis_mode)
    peng_rows = corr_peng.get("rows", [])

    for idx, r in enumerate(peng_rows, 1):
        comp_name = r.get("Komponen Pengeluaran", "-")
        pt = r.get("Tipe PDRB", "HK")
        rp = r.get("Korelasi dgn Penumpang")
        r2p = (rp ** 2) if (rp is not None and not np.isnan(rp)) else None
        pp = r.get("p-value Penumpang")

        rbag = r.get("Korelasi dgn Bagasi")
        r2bag = (rbag ** 2) if (rbag is not None and not np.isnan(rbag)) else None
        pbag = r.get("p-value Bagasi")

        rbar = r.get("Korelasi dgn Barang")
        r2bar = (rbar ** 2) if (rbar is not None and not np.isnan(rbar)) else None
        pbar = r.get("p-value Barang")

        fill_row = zebra_light_fill if idx % 2 == 0 else white_fill
        row_vals = [idx, comp_name, pt, rp, r2p, pp, rbag, r2bag, pbag, rbar, r2bar, pbar]

        for c_idx, v in enumerate(row_vals, 1):
            cell = ws1.cell(row=curr_row, column=c_idx, value=v)
            cell.font = data_font
            cell.fill = fill_row
            cell.border = thin_border
            if c_idx == 1:
                cell.alignment = align_center
            elif c_idx in [2, 3]:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                if c_idx in [4, 7, 10]:
                    cell.number_format = "0.0000"
                elif c_idx in [5, 8, 11]:
                    cell.number_format = "0.0%"
                elif c_idx in [6, 9, 12]:
                    cell.number_format = "0.0000"
        ws1.row_dimensions[curr_row].height = 18
        curr_row += 1

    autofit_columns(ws1)

    # =========================================================================
    # SHEET 2: 2_Diagnostik_VIF
    # =========================================================================
    ws2 = wb.create_sheet(title="2_Diagnostik_VIF")
    ws2.views.sheetView[0].showGridLines = True

    multi_res = compute_multicollinearity(province=province, price_type="HK", analysis_mode=analysis_mode, category="lu")
    cond_num = multi_res.get("condition_number", 1.0)
    method_str = multi_res.get("method", "Standard Inversion OLS")
    has_multi = multi_res.get("has_multicollinearity", False)
    status_overall = "⚠️ Terindikasi Multikolinearitas (Condition Number > 30 / VIF ≥ 5)" if has_multi else "✅ Bebas Multikolinearitas (Matrix Sehat & VIF < 5)"

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "DIAGNOSTIK MULTIKOLINEARITAS & VARIANCE INFLATION FACTOR (VIF)"
    ws2["A1"].font = title_font
    ws2["A1"].fill = navy_title_fill
    ws2["A1"].alignment = align_center
    ws2.row_dimensions[1].height = 28

    ws2.merge_cells("A2:G2")
    ws2["A2"] = f"Provinsi: {prov_a_name} | Mode: {analysis_label} | Condition Number (κ): {cond_num:.2f} | Status: {status_overall}"
    ws2["A2"].font = subtitle_font
    ws2["A2"].fill = navy_header_fill
    ws2["A2"].alignment = align_center
    ws2.row_dimensions[2].height = 18

    # Tabel 1: VIF 6 Sektor Inti
    curr_row = 4
    ws2.merge_cells(f"A{curr_row}:E{curr_row}")
    ws2[f"A{curr_row}"] = f"TABEL 1: VARIANCE INFLATION FACTOR (VIF) & TOLERANCE 6 SEKTOR INTI ({method_str})"
    ws2[f"A{curr_row}"].font = section_font
    ws2[f"A{curr_row}"].fill = indigo_section_fill
    ws2[f"A{curr_row}"].alignment = align_left
    ws2.row_dimensions[curr_row].height = 22
    curr_row += 1

    headers_vif = ["No", "Sektor PDRB (Subset Inti)", "Variance Inflation Factor (VIF)", "Tolerance (1/VIF)", "Status Evaluasi"]
    for c_idx, h in enumerate(headers_vif, 1):
        cell = ws2.cell(row=curr_row, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align_center
        cell.border = header_border
    ws2.row_dimensions[curr_row].height = 24
    curr_row += 1

    vif_results = multi_res.get("vif_results", [])
    for idx, v in enumerate(vif_results, 1):
        sec = v.get("sector", "-")
        vif_val = v.get("vif", 1.0)
        tol_val = v.get("tolerance", 1.0)
        st = v.get("status", "Aman")

        fill_row = zebra_light_fill if idx % 2 == 0 else white_fill
        c1 = ws2.cell(row=curr_row, column=1, value=idx)
        c2 = ws2.cell(row=curr_row, column=2, value=sec)
        c3 = ws2.cell(row=curr_row, column=3, value=vif_val)
        c4 = ws2.cell(row=curr_row, column=4, value=tol_val)
        c5 = ws2.cell(row=curr_row, column=5, value=st)

        c1.alignment = align_center
        c2.alignment = align_left
        c3.alignment = align_right
        c3.number_format = "0.00"
        c4.alignment = align_right
        c4.number_format = "0.0000"
        c5.alignment = align_center

        for cell in [c1, c2, c3, c4, c5]:
            cell.font = data_font
            cell.fill = fill_row
            cell.border = thin_border
        ws2.row_dimensions[curr_row].height = 18
        curr_row += 1

    # Tabel 2: Matriks Korelasi Antar-Sektor Rx
    curr_row += 1
    inter_mat = multi_res.get("inter_sector_matrix", {})
    mat_labels = inter_mat.get("labels", [])
    mat_vals = inter_mat.get("matrix", [])
    k_len = len(mat_labels)
    end_col_letter = get_column_letter(max(k_len + 1, 5))

    ws2.merge_cells(f"A{curr_row}:{end_col_letter}{curr_row}")
    ws2[f"A{curr_row}"] = "TABEL 2: MATRIKS KORELASI ANTAR-SEKTOR PDRB (Rx) [k x k]"
    ws2[f"A{curr_row}"].font = section_font
    ws2[f"A{curr_row}"].fill = indigo_section_fill
    ws2[f"A{curr_row}"].alignment = align_left
    ws2.row_dimensions[curr_row].height = 22
    curr_row += 1

    cell_h0 = ws2.cell(row=curr_row, column=1, value="Sektor PDRB")
    cell_h0.font = header_font
    cell_h0.fill = navy_header_fill
    cell_h0.alignment = align_center
    cell_h0.border = header_border

    for j, lbl in enumerate(mat_labels, 2):
        cell_hj = ws2.cell(row=curr_row, column=j, value=lbl)
        cell_hj.font = header_font
        cell_hj.fill = navy_header_fill
        cell_hj.alignment = align_center
        cell_hj.border = header_border
    ws2.row_dimensions[curr_row].height = 24
    curr_row += 1

    for i, row_lbl in enumerate(mat_labels):
        fill_row = zebra_light_fill if i % 2 == 0 else white_fill
        cell_r = ws2.cell(row=curr_row, column=1, value=row_lbl)
        cell_r.font = data_bold_font
        cell_r.fill = fill_row
        cell_r.border = thin_border
        cell_r.alignment = align_left

        for j in range(k_len):
            val = mat_vals[i][j] if i < len(mat_vals) and j < len(mat_vals[i]) else 0.0
            cell_v = ws2.cell(row=curr_row, column=j + 2, value=val)
            cell_v.font = data_font
            cell_v.fill = fill_row
            cell_v.border = thin_border
            cell_v.alignment = align_right
            cell_v.number_format = "0.0000"
        ws2.row_dimensions[curr_row].height = 18
        curr_row += 1

    # Tabel 3: Multikolinearitas Transportasi
    curr_row += 1
    trans_multi = multi_res.get("transport_multicollinearity", {})
    t_labels = trans_multi.get("labels", ["Penumpang", "Bagasi", "Barang"])
    t_mat = trans_multi.get("correlation_matrix", [])
    t_vifs = trans_multi.get("vif_results", [])

    ws2.merge_cells(f"A{curr_row}:E{curr_row}")
    ws2[f"A{curr_row}"] = "TABEL 3: MULTIKOLINEARITAS ANTAR-INDIKATOR TRANSPORTASI (RT) & VIF"
    ws2[f"A{curr_row}"].font = section_font
    ws2[f"A{curr_row}"].fill = indigo_section_fill
    ws2[f"A{curr_row}"].alignment = align_left
    ws2.row_dimensions[curr_row].height = 22
    curr_row += 1

    cell_t0 = ws2.cell(row=curr_row, column=1, value="Moda Transportasi")
    cell_t0.font = header_font
    cell_t0.fill = navy_header_fill
    cell_t0.alignment = align_center
    cell_t0.border = header_border

    for j, tlbl in enumerate(t_labels, 2):
        cell_tj = ws2.cell(row=curr_row, column=j, value=tlbl)
        cell_tj.font = header_font
        cell_tj.fill = navy_header_fill
        cell_tj.alignment = align_center
        cell_tj.border = header_border
    ws2.row_dimensions[curr_row].height = 24
    curr_row += 1

    for i, t_row_lbl in enumerate(t_labels):
        fill_row = zebra_light_fill if i % 2 == 0 else white_fill
        cell_tr = ws2.cell(row=curr_row, column=1, value=t_row_lbl)
        cell_tr.font = data_bold_font
        cell_tr.fill = fill_row
        cell_tr.border = thin_border
        cell_tr.alignment = align_left

        for j in range(len(t_labels)):
            val = t_mat[i][j] if i < len(t_mat) and j < len(t_mat[i]) else 0.0
            cell_tv = ws2.cell(row=curr_row, column=j + 2, value=val)
            cell_tv.font = data_font
            cell_tv.fill = fill_row
            cell_tv.border = thin_border
            cell_tv.alignment = align_right
            cell_tv.number_format = "0.0000"
        ws2.row_dimensions[curr_row].height = 18
        curr_row += 1

    curr_row += 1
    headers_tvif = ["No", "Indikator Transportasi", "VIF Transport", "Tolerance", "Status Multikolinearitas"]
    for c_idx, h in enumerate(headers_tvif, 1):
        cell = ws2.cell(row=curr_row, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align_center
        cell.border = header_border
    ws2.row_dimensions[curr_row].height = 24
    curr_row += 1

    for idx, tv in enumerate(t_vifs, 1):
        fill_row = zebra_light_fill if idx % 2 == 0 else white_fill
        c1 = ws2.cell(row=curr_row, column=1, value=idx)
        c2 = ws2.cell(row=curr_row, column=2, value=tv.get("indicator", "-"))
        c3 = ws2.cell(row=curr_row, column=3, value=tv.get("vif", 1.0))
        c4 = ws2.cell(row=curr_row, column=4, value=tv.get("tolerance", 1.0))
        c5 = ws2.cell(row=curr_row, column=5, value=tv.get("status", "Aman"))

        c1.alignment = align_center
        c2.alignment = align_left
        c3.alignment = align_right
        c3.number_format = "0.00"
        c4.alignment = align_right
        c4.number_format = "0.0000"
        c5.alignment = align_center

        for cell in [c1, c2, c3, c4, c5]:
            cell.font = data_font
            cell.fill = fill_row
            cell.border = thin_border
        ws2.row_dimensions[curr_row].height = 18
        curr_row += 1

    autofit_columns(ws2)

    # =========================================================================
    # SHEET 3: 3_Komparasi_Wilayah
    # =========================================================================
    ws3 = wb.create_sheet(title="3_Komparasi_Wilayah")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells("A1:I1")
    ws3["A1"] = f"KOMPARASI REGIONAL HEAD-TO-HEAD: {prov_a_name.upper()} VS {prov_b_name.upper()}"
    ws3["A1"].font = title_font
    ws3["A1"].fill = navy_title_fill
    ws3["A1"].alignment = align_center
    ws3.row_dimensions[1].height = 28

    ws3.merge_cells("A2:I2")
    ws3["A2"] = f"Indikator Transportasi: {trans_label} | Mode Analisis: {analysis_label} | Tahun: {year}"
    ws3["A2"].font = subtitle_font
    ws3["A2"].fill = navy_header_fill
    ws3["A2"].alignment = align_center
    ws3.row_dimensions[2].height = 18

    curr_row = 4
    headers_comp = [
        "No",
        "Sektor Lapangan Usaha",
        f"r Prov A ({prov_a_name})",
        f"R² Prov A (%)",
        f"r Prov B ({prov_b_name})",
        f"R² Prov B (%)",
        "Selisih Korelasi (|Δr|)",
        f"VIF ({prov_a_name})",
        f"VIF ({prov_b_name})"
    ]
    for c_idx, h in enumerate(headers_comp, 1):
        cell = ws3.cell(row=curr_row, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = amber_section_fill if "Selisih" in h else navy_header_fill
        cell.alignment = align_center
        cell.border = header_border
    ws3.row_dimensions[curr_row].height = 26
    curr_row += 1

    corr_a = _compute_correlations(province, year, "lu", "all", "default", "all", "", analysis_mode)
    corr_b = _compute_correlations(compare_province, year, "lu", "all", "default", "all", "", analysis_mode)

    rows_a = corr_a.get("rows", [])
    rows_b = corr_b.get("rows", [])

    map_a = {}
    for r in rows_a:
        sec = r.get("Lapangan Usaha", "")
        if "produk domestik" not in sec.lower() and "pdrb" not in sec.lower():
            map_a[_normalize_name(sec)] = (sec, r)

    map_b = {}
    for r in rows_b:
        sec = r.get("Lapangan Usaha", "")
        if "produk domestik" not in sec.lower() and "pdrb" not in sec.lower():
            map_b[_normalize_name(sec)] = (sec, r)

    multi_a = compute_multicollinearity(province, "HK", analysis_mode, "lu")
    multi_b = compute_multicollinearity(compare_province, "HK", analysis_mode, "lu")

    vif_dict_a = {_normalize_name(v["sector"]): v["vif"] for v in multi_a.get("vif_results", [])}
    vif_dict_b = {_normalize_name(v["sector"]): v["vif"] for v in multi_b.get("vif_results", [])}

    target_metric_col = "Korelasi dgn Penumpang"
    if transport_type.lower() == "bagasi":
        target_metric_col = "Korelasi dgn Bagasi"
    elif transport_type.lower() == "barang":
        target_metric_col = "Korelasi dgn Barang"

    all_keys = list(map_a.keys())
    for k in map_b.keys():
        if k not in all_keys:
            all_keys.append(k)

    comp_table_rows = []
    for k in all_keys:
        sec_name = map_a.get(k, (map_b.get(k, ("", {}))[0], {}))[0]
        row_a = map_a.get(k, ("", {}))[1]
        row_b = map_b.get(k, ("", {}))[1]

        r_a = row_a.get(target_metric_col)
        r2_a = (r_a ** 2) if (r_a is not None and not np.isnan(r_a)) else None

        r_b = row_b.get(target_metric_col)
        r2_b = (r_b ** 2) if (r_b is not None and not np.isnan(r_b)) else None

        if r_a is not None and r_b is not None and not np.isnan(r_a) and not np.isnan(r_b):
            delta = abs(r_a - r_b)
        else:
            delta = None

        vif_a_val = vif_dict_a.get(k)
        vif_b_val = vif_dict_b.get(k)

        comp_table_rows.append({
            "sector": sec_name,
            "r_a": r_a,
            "r2_a": r2_a,
            "r_b": r_b,
            "r2_b": r2_b,
            "delta": delta,
            "vif_a": vif_a_val,
            "vif_b": vif_b_val
        })

    comp_table_rows.sort(key=lambda x: (x["delta"] is not None, x["delta"] or 0), reverse=True)

    for idx, item in enumerate(comp_table_rows, 1):
        fill_row = zebra_light_fill if idx % 2 == 0 else white_fill
        c1 = ws3.cell(row=curr_row, column=1, value=idx)
        c2 = ws3.cell(row=curr_row, column=2, value=item["sector"])
        c3 = ws3.cell(row=curr_row, column=3, value=item["r_a"])
        c4 = ws3.cell(row=curr_row, column=4, value=item["r2_a"])
        c5 = ws3.cell(row=curr_row, column=5, value=item["r_b"])
        c6 = ws3.cell(row=curr_row, column=6, value=item["r2_b"])
        c7 = ws3.cell(row=curr_row, column=7, value=item["delta"])
        c8 = ws3.cell(row=curr_row, column=8, value=item["vif_a"])
        c9 = ws3.cell(row=curr_row, column=9, value=item["vif_b"])

        c1.alignment = align_center
        c2.alignment = align_left
        c3.alignment = align_right
        c3.number_format = "0.0000"
        c4.alignment = align_right
        c4.number_format = "0.0%"
        c5.alignment = align_right
        c5.number_format = "0.0000"
        c6.alignment = align_right
        c6.number_format = "0.0%"
        c7.alignment = align_right
        c7.number_format = "0.0000"
        c8.alignment = align_right
        c8.number_format = "0.00" if item["vif_a"] is not None else "@"
        c9.alignment = align_right
        c9.number_format = "0.00" if item["vif_b"] is not None else "@"

        for cell in [c1, c2, c3, c4, c5, c6, c7, c8, c9]:
            cell.font = data_font
            cell.fill = fill_row
            cell.border = thin_border

        ws3.row_dimensions[curr_row].height = 18
        curr_row += 1

    autofit_columns(ws3)

    # =========================================================================
    # SHEET 4: 4_Tren_Triwulanan
    # =========================================================================
    ws4 = wb.create_sheet(title="4_Tren_Triwulanan")
    ws4.views.sheetView[0].showGridLines = True

    ws4.merge_cells("A1:O1")
    ws4["A1"] = "DERET WAKTU TRIWULANAN (2020 - 2024) PDRB & TRANSPORTASI UDARA"
    ws4["A1"].font = title_font
    ws4["A1"].fill = navy_title_fill
    ws4["A1"].alignment = align_center
    ws4.row_dimensions[1].height = 28

    ws4.merge_cells("A2:O2")
    ws4["A2"] = f"Provinsi: {prov_a_name} | Rentang Runtun Waktu: 2020 Q1 s/d 2024 Q4 | BPS RI"
    ws4["A2"].font = subtitle_font
    ws4["A2"].fill = navy_header_fill
    ws4["A2"].alignment = align_center
    ws4.row_dimensions[2].height = 18

    curr_row = 4
    headers_ts = [
        "Tahun", "Triwulan",
        "Penumpang (Orang)", "Bagasi (Kg)", "Barang (Kg)",
        "PDRB Total HK (Juta Rp)", "PDRB Total HB (Juta Rp)",
        "YoY Penumpang (%)", "YoY Bagasi (%)", "YoY Barang (%)", "YoY PDRB Total HK (%)",
        "QoQ Penumpang (%)", "QoQ Bagasi (%)", "QoQ Barang (%)", "QoQ PDRB Total HK (%)"
    ]
    for c_idx, h in enumerate(headers_ts, 1):
        cell = ws4.cell(row=curr_row, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align_center
        cell.border = header_border
    ws4.row_dimensions[curr_row].height = 26
    curr_row += 1

    df_multi, _ = _build_multi_year_df_cached(province)
    df_yoy = _get_growth_df_cached(province, 4)
    df_qoq = _get_growth_df_cached(province, 1)

    p_col = next((c for c in df_multi.columns if "penumpang" in c.lower()), None)
    bag_col = next((c for c in df_multi.columns if "bagasi" in c.lower()), None)
    bar_col = next((c for c in df_multi.columns if "barang" in c.lower()), None)

    tot_hk_col = next((c for c in df_multi.columns if c.startswith("LU (HK)") and ("produk domestik" in c.lower() or "pdrb" in c.lower())), None)
    tot_hb_col = next((c for c in df_multi.columns if c.startswith("LU (HB)") and ("produk domestik" in c.lower() or "pdrb" in c.lower())), None)

    for idx, r in df_multi.iterrows():
        yr_val = int(r["_tahun"])
        tw_num = int(r["_triwulan_num"])
        tw_str = f"Triwulan {tw_num}"

        val_p = float(r[p_col]) if p_col and pd.notna(r[p_col]) else 0
        val_bag = float(r[bag_col]) if bag_col and pd.notna(r[bag_col]) else 0
        val_bar = float(r[bar_col]) if bar_col and pd.notna(r[bar_col]) else 0
        val_hk = float(r[tot_hk_col]) if tot_hk_col and pd.notna(r[tot_hk_col]) else 0
        val_hb = float(r[tot_hb_col]) if tot_hb_col and pd.notna(r[tot_hb_col]) else 0

        yoy_p = None
        yoy_bag = None
        yoy_bar = None
        yoy_hk = None
        if not df_yoy.empty:
            match_yoy = df_yoy[(df_yoy["_tahun"] == yr_val) & (df_yoy["_triwulan_num"] == tw_num)]
            if not match_yoy.empty:
                m_row = match_yoy.iloc[0]
                yoy_p = float(m_row[p_col]) if p_col and pd.notna(m_row.get(p_col)) else None
                yoy_bag = float(m_row[bag_col]) if bag_col and pd.notna(m_row.get(bag_col)) else None
                yoy_bar = float(m_row[bar_col]) if bar_col and pd.notna(m_row.get(bar_col)) else None
                yoy_hk = float(m_row[tot_hk_col]) if tot_hk_col and pd.notna(m_row.get(tot_hk_col)) else None

        qoq_p = None
        qoq_bag = None
        qoq_bar = None
        qoq_hk = None
        if not df_qoq.empty:
            match_qoq = df_qoq[(df_qoq["_tahun"] == yr_val) & (df_qoq["_triwulan_num"] == tw_num)]
            if not match_qoq.empty:
                m_row = match_qoq.iloc[0]
                qoq_p = float(m_row[p_col]) if p_col and pd.notna(m_row.get(p_col)) else None
                qoq_bag = float(m_row[bag_col]) if bag_col and pd.notna(m_row.get(bag_col)) else None
                qoq_bar = float(m_row[bar_col]) if bar_col and pd.notna(m_row.get(bar_col)) else None
                qoq_hk = float(m_row[tot_hk_col]) if tot_hk_col and pd.notna(m_row.get(tot_hk_col)) else None

        fill_row = zebra_light_fill if curr_row % 2 == 0 else white_fill
        ts_row_vals = [
            yr_val, tw_str,
            val_p, val_bag, val_bar, val_hk, val_hb,
            (yoy_p / 100.0) if yoy_p is not None else None,
            (yoy_bag / 100.0) if yoy_bag is not None else None,
            (yoy_bar / 100.0) if yoy_bar is not None else None,
            (yoy_hk / 100.0) if yoy_hk is not None else None,
            (qoq_p / 100.0) if qoq_p is not None else None,
            (qoq_bag / 100.0) if qoq_bag is not None else None,
            (qoq_bar / 100.0) if qoq_bar is not None else None,
            (qoq_hk / 100.0) if qoq_hk is not None else None
        ]

        for c_idx, v in enumerate(ts_row_vals, 1):
            cell = ws4.cell(row=curr_row, column=c_idx, value=v)
            cell.font = data_font
            cell.fill = fill_row
            cell.border = thin_border
            if c_idx in [1, 2]:
                cell.alignment = align_center
            elif c_idx in [3, 4, 5, 6, 7]:
                cell.alignment = align_right
                cell.number_format = "#,##0" if c_idx in [3, 4, 5] else "#,##0.00"
            else:
                cell.alignment = align_right
                cell.number_format = "0.0%" if v is not None else "@"

        ws4.row_dimensions[curr_row].height = 18
        curr_row += 1

    autofit_columns(ws4)

    # =========================================================================
    # SHEET 5: 5_Data_Mentah
    # =========================================================================
    ws5 = wb.create_sheet(title="5_Data_Mentah")
    ws5.views.sheetView[0].showGridLines = True

    lu_hk_cols = [c for c in df_multi.columns if c.startswith("LU (HK)")]
    lu_hb_cols = [c for c in df_multi.columns if c.startswith("LU (HB)")]

    max_c5 = max(len(lu_hk_cols) + 5, 10)
    end_col_5 = get_column_letter(max_c5)

    ws5.merge_cells(f"A1:{end_col_5}1")
    ws5["A1"] = "DATA MENTAH 17 SEKTOR PDRB LAPANGAN USAHA (ADHK & ADHB) & TRANSPORTASI"
    ws5["A1"].font = title_font
    ws5["A1"].fill = navy_title_fill
    ws5["A1"].alignment = align_center
    ws5.row_dimensions[1].height = 28

    ws5.merge_cells(f"A2:{end_col_5}2")
    ws5["A2"] = f"Provinsi: {prov_a_name} | Satuan: Juta Rupiah (PDRB) / Satuan Riil Transportasi | 2020-2024"
    ws5["A2"].font = subtitle_font
    ws5["A2"].fill = navy_header_fill
    ws5["A2"].alignment = align_center
    ws5.row_dimensions[2].height = 18

    # Tabel 1: PDRB Harga Konstan (ADHK)
    curr_row = 4
    ws5.merge_cells(f"A{curr_row}:{end_col_5}{curr_row}")
    ws5[f"A{curr_row}"] = "TABEL 1: PDRB ATAS DASAR HARGA KONSTAN (ADHK 2010) MENURUT 17 LAPANGAN USAHA & TRANSPORTASI"
    ws5[f"A{curr_row}"].font = section_font
    ws5[f"A{curr_row}"].fill = emerald_section_fill
    ws5[f"A{curr_row}"].alignment = align_left
    ws5.row_dimensions[curr_row].height = 22
    curr_row += 1

    headers_adhk = ["Tahun", "Triwulan"] + [re.sub(r'^LU\s*\(HK\)\s*-\s*', '', c).strip() for c in lu_hk_cols] + ["Penumpang (Orang)", "Bagasi (Kg)", "Barang (Kg)"]
    for c_idx, h in enumerate(headers_adhk, 1):
        cell = ws5.cell(row=curr_row, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align_center
        cell.border = header_border
    ws5.row_dimensions[curr_row].height = 24
    curr_row += 1

    for idx, r in df_multi.iterrows():
        yr_val = int(r["_tahun"])
        tw_str = f"Triwulan {int(r['_triwulan_num'])}"
        vals_hk = [float(r[c]) if pd.notna(r[c]) else 0.0 for c in lu_hk_cols]
        val_p = float(r[p_col]) if p_col and pd.notna(r[p_col]) else 0.0
        val_bag = float(r[bag_col]) if bag_col and pd.notna(r[bag_col]) else 0.0
        val_bar = float(r[bar_col]) if bar_col and pd.notna(r[bar_col]) else 0.0

        fill_row = zebra_light_fill if curr_row % 2 == 0 else white_fill
        row_cells = [yr_val, tw_str] + vals_hk + [val_p, val_bag, val_bar]

        for c_idx, v in enumerate(row_cells, 1):
            cell = ws5.cell(row=curr_row, column=c_idx, value=v)
            cell.font = data_font
            cell.fill = fill_row
            cell.border = thin_border
            if c_idx in [1, 2]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
                cell.number_format = "#,##0.00" if c_idx <= len(lu_hk_cols) + 2 else "#,##0"
        ws5.row_dimensions[curr_row].height = 18
        curr_row += 1

    # Tabel 2: PDRB Harga Berlaku (ADHB)
    curr_row += 2
    ws5.merge_cells(f"A{curr_row}:{end_col_5}{curr_row}")
    ws5[f"A{curr_row}"] = "TABEL 2: PDRB ATAS DASAR HARGA BERLAKU (ADHB NOMINAL) MENURUT 17 LAPANGAN USAHA"
    ws5[f"A{curr_row}"].font = section_font
    ws5[f"A{curr_row}"].fill = emerald_section_fill
    ws5[f"A{curr_row}"].alignment = align_left
    ws5.row_dimensions[curr_row].height = 22
    curr_row += 1

    headers_adhb = ["Tahun", "Triwulan"] + [re.sub(r'^LU\s*\(HB\)\s*-\s*', '', c).strip() for c in lu_hb_cols]
    for c_idx, h in enumerate(headers_adhb, 1):
        cell = ws5.cell(row=curr_row, column=c_idx, value=h)
        cell.font = header_font
        cell.fill = navy_header_fill
        cell.alignment = align_center
        cell.border = header_border
    ws5.row_dimensions[curr_row].height = 24
    curr_row += 1

    for idx, r in df_multi.iterrows():
        yr_val = int(r["_tahun"])
        tw_str = f"Triwulan {int(r['_triwulan_num'])}"
        vals_hb = [float(r[c]) if pd.notna(r[c]) else 0.0 for c in lu_hb_cols]

        fill_row = zebra_light_fill if curr_row % 2 == 0 else white_fill
        row_cells = [yr_val, tw_str] + vals_hb

        for c_idx, v in enumerate(row_cells, 1):
            cell = ws5.cell(row=curr_row, column=c_idx, value=v)
            cell.font = data_font
            cell.fill = fill_row
            cell.border = thin_border
            if c_idx in [1, 2]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right
                cell.number_format = "#,##0.00"
        ws5.row_dimensions[curr_row].height = 18
        curr_row += 1

    autofit_columns(ws5)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.get("/api/export-full-excel")
async def get_export_full_excel(
    province: str = "sulawesi_selatan",
    compare_province: str = "gorontalo",
    year: str = "2024",
    analysis_mode: str = "growth_yoy",
    transport_type: str = "penumpang"
):
    """
    Endpoint GET untuk menghasilkan dan mengunduh buku kerja Excel (.xlsx) 5-sheet
    berdasarkan parameter analisis aktif.
    """
    try:
        buf = generate_full_excel_report(
            province=province,
            compare_province=compare_province,
            year=year,
            analysis_mode=analysis_mode,
            transport_type=transport_type
        )
        filename = f"Laporan_Analisis_PDRB_{province}_{year}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Cache-Control": "no-cache"
            }
        )
    except Exception as e:
        import traceback
        print(f"⚠️ Error generating Excel report: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Gagal menghasilkan laporan Excel: {str(e)}")


@app.post("/api/export-full-excel")
async def post_export_full_excel(req: ExportExcelRequest):
    """
    Endpoint POST untuk menghasilkan dan mengunduh buku kerja Excel (.xlsx) 5-sheet
    melalui body JSON.
    """
    try:
        buf = generate_full_excel_report(
            province=req.province,
            compare_province=req.compare_province,
            year=req.year,
            analysis_mode=req.analysis_mode,
            transport_type=req.transport_type
        )
        filename = f"Laporan_Analisis_PDRB_{req.province}_{req.year}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Cache-Control": "no-cache"
            }
        )
    except Exception as e:
        import traceback
        print(f"⚠️ Error generating Excel report: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Gagal menghasilkan laporan Excel: {str(e)}")


# ======================================================================================
# STARTUP CACHE PRE-WARMING
# ======================================================================================

def prewarm_all_caches():
    """Melakukan pra-kalkulasi seluruh dataset di memory saat server pertama kali start."""
    try:
        manifest = load_manifest()
        prov_list = list(manifest.get("provinces", {}).keys())
        for p in prov_list:
            _build_multi_year_df_cached(p)
            _get_growth_df_cached(p, 4)
            _get_growth_df_cached(p, 1)
            for cat in ["lu", "peng"]:
                for pt in ["HK", "HB", "ALL"]:
                    _get_growth_correlation_matrix_cached(p, 4, cat, pt)
                    _get_growth_correlation_matrix_cached(p, 1, cat, pt)
            # Prewarm default multicollinearity
            compute_multicollinearity(province=p, price_type="HK", analysis_mode="growth_yoy", category="lu")
        try:
            print(f"⚡ In-memory cache pre-warmed for {len(prov_list)} provinces.")
        except Exception:
            pass
    except Exception as e:
        try:
            print(f"⚠️ Cache pre-warming note: {e}")
        except Exception:
            pass


# Jalankan pre-warming saat startup aplikasi
@app.on_event("startup")
def on_startup():
    prewarm_all_caches()


# Jalankan synchronous warmup untuk lingkungan serverless cold-start
prewarm_all_caches()

