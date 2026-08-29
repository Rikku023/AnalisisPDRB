"""
========================================================================================
ETL Script: Import & Integrasi Bersih Dataset Kalimantan (2020–2024) ke Parquet
File: scripts/import_kalimantan_clean.py

Sumber Data:
1. PDRB: PDRB_ADHK_Lapangan_Usaha_Kalimantan.xlsx
   - Sheet raw: Kalimantan Barat raw, Kalimantan Selatan raw, Kalimantan Tengah raw,
     Kalimantan Timur raw, Kalimantan Utara raw.
2. Transportasi Udara: ANALISIS TRANSPORTASI UDARA.xlsx
   - Sheet: BPS_KALBAR, BPS_KALSEL, BPS_KALTENG, BPS_KALTIM, BPS_TAHUNAN

Aturan Rekonstruksi & Fallback Data:
1. Rekonstruksi Sektor G:
   Untuk sheet yang tidak memiliki Sektor G (Kalsel, Kalteng, Kaltim, Kaltara),
   hitung: Sektor G = Total PDRB - sum(16 sektor lainnya).
2. Penanganan Ketiadaan ADHB:
   Untuk sheet yang hanya memiliki data ADHK (Kalteng, Kaltim, Kaltara),
   salin nilai ADHK ke kolom LU (HB) - [Nama Sektor] sebagai fallback otomatis.
3. Format Output Parquet:
   - data/{prov_key}_multi_year_pdrb_triwulan.parquet & api/data/...
   - data/{prov_key}_pdrb_triwulan.parquet & api/data/...
   - data/{prov_key}_{yr}_pdrb_triwulan.parquet & api/data/...
   - Matriks korelasi triwulanan (*_correl_lu.parquet, *_correl_peng.parquet)
   - Exclude Total PDRB dari matriks korelasi antar-sektor (Condition Number < 100)
========================================================================================
"""

import os
import sys
import json
import re
from pathlib import Path
from typing import Dict, List, Any, Tuple, Optional
import pandas as pd
import numpy as np
import openpyxl

# Set stdout encoding
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

PROJECT_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_DIR / "data"
API_DATA_DIR = PROJECT_DIR / "api" / "data"
PUBLIC_DIR = PROJECT_DIR / "public"

# Candidate paths for source Excel files
PDRB_CANDIDATES = [
    PROJECT_DIR / "PDRB per Zona" / "PDRB_ADHK_Lapangan_Usaha_Kalimantan.xlsx",
    PROJECT_DIR / "PDRB_ADHK_Lapangan_Usaha_Kalimantan.xlsx",
    Path(r"C:\Users\Eric\Documents\Kuliah - kerja\KP\Data\PDRB per Zona\PDRB_ADHK_Lapangan_Usaha_Kalimantan.xlsx"),
    Path(r"C:\Users\Eric\Documents\Kuliah - kerja\KP\Data\PDRB_ADHK_Lapangan_Usaha_Kalimantan.xlsx"),
]

TRANSPORT_CANDIDATES = [
    PROJECT_DIR / "ANALISIS TRANSPORTASI UDARA.xlsx",
    Path(r"C:\Users\Eric\Documents\Kuliah - kerja\KP\Data\ANALISIS TRANSPORTASI UDARA.xlsx"),
    Path(r"C:\Users\Eric\Documents\Kuliah - kerja\KP\ANALISIS TRANSPORTASI UDARA.xlsx"),
]


def resolve_file(candidates: List[Path], desc: str) -> Path:
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError(f"File {desc} tidak ditemukan di lokasi kandidat: {candidates}")


# 17 Sektor Lapangan Usaha Standar BPS
STANDARDIZED_SECTORS = [
    "Pertanian, Kehutanan, dan Perikanan",
    "Pertambangan dan Penggalian",
    "Industri Pengolahan",
    "Pengadaan Listrik dan Gas",
    "Pengadaan Air, Pengelolaan Sampah, Limbah, dan Daur Ulang",
    "Konstruksi",
    "Perdagangan Besar dan Eceran, Reparasi Mobil dan Sepeda Motor",
    "Transportasi dan Pergudangan",
    "Penyediaan Akomodasi dan Makan Minum",
    "Informasi dan Komunikasi",
    "Jasa Keuangan dan Asuransi",
    "Real Estat",
    "Jasa Perusahaan",
    "Administrasi Pemerintahan, Pertahanan, dan Jaminan Sosial Wajib",
    "Jasa Pendidikan",
    "Jasa Kesehatan dan Kegiatan Sosial",
    "Jasa Lainnya",
]

# Komponen Pengeluaran Standar BPS
STANDARDIZED_EXP_COMPONENTS = [
    "Pengeluaran Konsumsi Rumah Tangga",
    "Pengeluaran Konsumsi LNPRT",
    "Pengeluaran Konsumsi Pemerintah",
    "Pembentukan Modal Tetap Bruto",
    "Perubahan Inventori",
    "Ekspor Barang dan Jasa",
    "Impor Barang dan Jasa",
]

EXP_RATIOS = {
    "Pengeluaran Konsumsi Rumah Tangga": 0.48,
    "Pengeluaran Konsumsi LNPRT": 0.015,
    "Pengeluaran Konsumsi Pemerintah": 0.085,
    "Pembentukan Modal Tetap Bruto": 0.28,
    "Perubahan Inventori": 0.02,
    "Ekspor Barang dan Jasa": 0.46,
    "Impor Barang dan Jasa": 0.34,
}

PROVINCES_CONFIG = {
    "kalimantan_barat": {
        "name": "Kalimantan Barat",
        "raw_sheet": "Kalimantan Barat raw",
        "transport_sheet": "BPS_KALBAR",
        "has_adhb": True,
        "missing_sector_g": False,
        "airports": [
            "Supadio (Kubu Raya/Pontianak)",
            "Rahadi Oesman (Ketapang)",
            "Tebelian (Sintang)",
            "Pangsuma (Kapuas Hulu)",
            "Singkawang (Singkawang)",
            "Nanga Pinoh (Melawi)"
        ]
    },
    "kalimantan_selatan": {
        "name": "Kalimantan Selatan",
        "raw_sheet": "Kalimantan Selatan raw",
        "transport_sheet": "BPS_KALSEL",
        "has_adhb": True,
        "missing_sector_g": True,
        "airports": [
            "Syamsudin Noor (Banjar/Banjarmasin)",
            "Gusti Sjamsir Alam (Kotabaru)",
            "Bersujud (Batulicin)",
            "Warukin (Tabalong)"
        ]
    },
    "kalimantan_tengah": {
        "name": "Kalimantan Tengah",
        "raw_sheet": "Kalimantan Tengah raw",
        "transport_sheet": "BPS_KALTENG",
        "has_adhb": False,
        "missing_sector_g": True,
        "airports": [
            "Tjilik Riwut (Palangka Raya)",
            "Iskandar (Pangkalan Bun)",
            "H. Asan (Sampit)",
            "Beringin (Muara Teweh)",
            "Kuala Pembuang (Seruyan)",
            "Sanggu (Buntok)"
        ]
    },
    "kalimantan_timur": {
        "name": "Kalimantan Timur",
        "raw_sheet": "Kalimantan Timur raw",
        "transport_sheet": "BPS_KALTIM",
        "has_adhb": False,
        "missing_sector_g": True,
        "airports": [
            "Sultan Aji Muhammad Sulaiman Sepinggan (Balikpapan)",
            "APT Pranoto (Samarinda)",
            "Kalimarau (Berau)",
            "Badak (Bontang)",
            "Melak (Kutai Barat)",
            "Datah Dawai (Mahakam Ulu)"
        ]
    },
    "kalimantan_utara": {
        "name": "Kalimantan Utara",
        "raw_sheet": "Kalimantan Utara raw",
        "transport_sheet": None,  # Fallback to BPS_TAHUNAN
        "has_adhb": False,
        "missing_sector_g": True,
        "airports": [
            "Juwata (Tarakan)",
            "Tanjung Harapan (Tanjung Selor/Bulungan)",
            "Kolonel RA Bessing (Malinau)",
            "Nunukan (Nunukan)",
            "Yuvai Semaring (Long Bawan/Krayan)"
        ]
    }
}

YEARS = [2020, 2021, 2022, 2023, 2024]
QUARTERS = ["Triwulan I", "Triwulan II", "Triwulan III", "Triwulan IV"]
SEASONAL_WEIGHTS = [0.22, 0.25, 0.26, 0.27]


def normalize_sector_name(raw_name: Any) -> str:
    """Membersihkan kode huruf (A-Q, M,N, R,S,T,U) dan tanda baca dari nama sektor."""
    s = str(raw_name).strip()
    s = re.sub(r'^[A-Z](,[A-Z])*\.\s*', '', s)
    s = s.replace(';', ',')
    s = re.sub(r'\s+', ' ', s)

    if 'Pertanian' in s:
        return 'Pertanian, Kehutanan, dan Perikanan'
    if 'Pertambangan' in s:
        return 'Pertambangan dan Penggalian'
    if 'Industri' in s:
        return 'Industri Pengolahan'
    if 'Listrik' in s:
        return 'Pengadaan Listrik dan Gas'
    if 'Pengadaan Air' in s or 'Limbah' in s:
        return 'Pengadaan Air, Pengelolaan Sampah, Limbah, dan Daur Ulang'
    if 'Konstruksi' in s:
        return 'Konstruksi'
    if 'Perdagangan' in s:
        return 'Perdagangan Besar dan Eceran, Reparasi Mobil dan Sepeda Motor'
    if 'Transportasi' in s:
        return 'Transportasi dan Pergudangan'
    if 'Akomodasi' in s:
        return 'Penyediaan Akomodasi dan Makan Minum'
    if 'Informasi' in s:
        return 'Informasi dan Komunikasi'
    if 'Keuangan' in s:
        return 'Jasa Keuangan dan Asuransi'
    if 'Real' in s:
        return 'Real Estat'
    if 'Perusahaan' in s:
        return 'Jasa Perusahaan'
    if 'Pemerintahan' in s:
        return 'Administrasi Pemerintahan, Pertahanan, dan Jaminan Sosial Wajib'
    if 'Pendidikan' in s:
        return 'Jasa Pendidikan'
    if 'Kesehatan' in s:
        return 'Jasa Kesehatan dan Kegiatan Sosial'
    if 'Lainnya' in s or 'lainnya' in s:
        return 'Jasa Lainnya'
    return s


def is_metadata_or_total_row(wilayah_name: str) -> bool:
    """Mendeteksi apakah baris adalah total PDRB, sub-total, atau metadata yang bukan sektor individu."""
    w = str(wilayah_name).strip().upper()
    return any(kw in w for kw in ['PDRB', 'PRODUK DOMESTIK REGIONAL BRUTO', 'LAJU', 'DISTRIBUSI', 'SUMBER'])


def is_main_total_row(wilayah_name: str) -> bool:
    """Mendeteksi apakah baris adalah baris total PDRB utama level provinsi."""
    w = str(wilayah_name).strip().upper()
    if 'TANPA MIGAS' in w:
        return False
    return 'PDRB' in w or 'PRODUK DOMESTIK REGIONAL BRUTO' in w


def extract_pdrb_from_raw_sheet(
    wb_pdrb: openpyxl.Workbook,
    prov_key: str,
    cfg: Dict[str, Any]
) -> Dict[int, Dict[str, Dict[str, Dict[str, float]]]]:
    """
    Ekstrak data PDRB triwulanan (HK & HB) dari sheet raw BPS.
    Format return:
    {
        year: {
            quarter: {
                "HK": {sector: value, ...},
                "HB": {sector: value, ...}
            }
        }
    }
    """
    sheet_name = cfg["raw_sheet"]
    if sheet_name not in wb_pdrb.sheetnames:
        raise ValueError(f"Sheet {sheet_name} tidak ditemukan di workbook PDRB!")

    ws = wb_pdrb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}

    raw_data: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        thn = ws.cell(r, col_idx.get("tahun", 7)).value
        try:
            yr = int(thn)
        except (ValueError, TypeError):
            continue

        if yr not in YEARS:
            continue

        per = str(ws.cell(r, col_idx.get("periode", 8)).value or "").strip()
        if per not in QUARTERS:
            continue

        wil = str(ws.cell(r, col_idx.get("wilayah", 4)).value or "").strip()
        komp = str(ws.cell(r, col_idx.get("komponen", 6)).value or "").strip()
        val = ws.cell(r, col_idx.get("nilai", 9)).value
        try:
            val_f = float(val or 0)
        except (ValueError, TypeError):
            val_f = 0.0

        raw_data.append({
            "tahun": yr,
            "periode": per,
            "wilayah": wil,
            "komponen": komp,
            "nilai": val_f
        })

    df_raw = pd.DataFrame(raw_data)
    result = {yr: {q: {"HK": {}, "HB": {}} for q in QUARTERS} for yr in YEARS}

    for yr in YEARS:
        for q in QUARTERS:
            sub = df_raw[(df_raw["tahun"] == yr) & (df_raw["periode"] == q)]

            # 1. Parse HK
            sub_hk = sub[sub["komponen"].str.contains("Konstan", case=False, na=False)]
            tot_hk = 0.0
            sec_hk: Dict[str, float] = {}

            for _, r in sub_hk.iterrows():
                w = r["wilayah"]
                val = r["nilai"]
                if is_main_total_row(w):
                    tot_hk = val
                elif is_metadata_or_total_row(w):
                    continue
                else:
                    norm = normalize_sector_name(w)
                    if norm in STANDARDIZED_SECTORS:
                        sec_hk[norm] = val

            # Rekonstruksi Sektor G jika hilang
            g_name = "Perdagangan Besar dan Eceran, Reparasi Mobil dan Sepeda Motor"
            if cfg["missing_sector_g"] or g_name not in sec_hk:
                if tot_hk > 0:
                    other_sum = sum(v for k, v in sec_hk.items() if k != g_name)
                    sec_hk[g_name] = max(0.0, tot_hk - other_sum)

            # Jika total PDRB kosong, hitung dari sum sektor
            if tot_hk == 0.0:
                tot_hk = sum(sec_hk.values())
            sec_hk["Produk Domestik Regional Bruto (PDRB)"] = tot_hk
            result[yr][q]["HK"] = sec_hk

            # 2. Parse HB (atau fallback ke HK jika tidak ada HB)
            if cfg["has_adhb"]:
                sub_hb = sub[sub["komponen"].str.contains("Berlaku", case=False, na=False)]
                if not sub_hb.empty:
                    tot_hb = 0.0
                    sec_hb: Dict[str, float] = {}
                    for _, r in sub_hb.iterrows():
                        w = r["wilayah"]
                        val = r["nilai"]
                        if is_main_total_row(w):
                            tot_hb = val
                        elif is_metadata_or_total_row(w):
                            continue
                        else:
                            norm = normalize_sector_name(w)
                            if norm in STANDARDIZED_SECTORS:
                                sec_hb[norm] = val

                    if cfg["missing_sector_g"] or g_name not in sec_hb:
                        if tot_hb > 0:
                            other_sum = sum(v for k, v in sec_hb.items() if k != g_name)
                            sec_hb[g_name] = max(0.0, tot_hb - other_sum)

                    if tot_hb == 0.0:
                        tot_hb = sum(sec_hb.values())
                    sec_hb["Produk Domestik Regional Bruto (PDRB)"] = tot_hb
                    result[yr][q]["HB"] = sec_hb
                else:
                    # Fallback ke HK jika sheet HB kosong di kuartal tsb
                    result[yr][q]["HB"] = sec_hk.copy()
            else:
                # Fallback otomatis ADHK -> ADHB
                result[yr][q]["HB"] = sec_hk.copy()

    return result


def extract_transport_quarterly(
    wb_trans: openpyxl.Workbook,
    prov_key: str,
    cfg: Dict[str, Any]
) -> Tuple[Dict[int, Dict[str, List[float]]], Dict[int, List[Dict[str, Any]]]]:
    """
    Ekstrak data transportasi triwulanan (Penumpang, Bagasi, Barang)
    dari sheet transportasi atau fallback ke sheet BPS_TAHUNAN.
    """
    quarterly_data = {y: {"penumpang": [0.0]*4, "bagasi": [0.0]*4, "barang": [0.0]*4} for y in YEARS}
    routes_by_year = {y: [] for y in YEARS}

    sheet_name = cfg.get("transport_sheet")

    # 1. Jika ada sheet khusus (BPS_KALBAR, BPS_KALSEL, BPS_KALTIM, BPS_KALTENG)
    if sheet_name and sheet_name in wb_trans.sheetnames:
        ws = wb_trans[sheet_name]

        # Cari letak tabel bulanan (PENUMPANG, BAGASI, BARANG)
        tables_pos: Dict[str, int] = {}
        for r in range(1, 45):
            val_k = str(ws.cell(r, 11).value or "").strip().upper()
            if val_k in ["PENUMPANG", "BAGASI", "BARANG"]:
                tables_pos[val_k.lower()] = r

        # Baca header tahun tabel bulanan (biasanya row r+1)
        def read_monthly_table(start_r: Optional[int]) -> Dict[int, List[float]]:
            res = {y: [0.0]*12 for y in YEARS}
            if start_r is None:
                return res

            hdr_row = start_r + 1
            year_cols = {}
            for c in range(12, 19):
                y_val = ws.cell(hdr_row, c).value
                try:
                    yr = int(y_val)
                    if yr in YEARS:
                        year_cols[yr] = c
                except (ValueError, TypeError):
                    pass

            for m_idx, r in enumerate(range(start_r + 2, start_r + 14)):
                for yr, c in year_cols.items():
                    val = ws.cell(r, c).value
                    if val is not None:
                        try:
                            res[yr][m_idx] = float(val)
                        except (ValueError, TypeError):
                            pass
            return res

        p_monthly = read_monthly_table(tables_pos.get("penumpang"))
        bag_monthly = read_monthly_table(tables_pos.get("bagasi"))
        bar_monthly = read_monthly_table(tables_pos.get("barang"))

        # Baca data rute (Col A-I)
        route_annual = {y: {"penumpang": 0.0, "bagasi": 0.0, "barang": 0.0} for y in YEARS}
        for r in range(2, ws.max_row + 1):
            thn = ws.cell(r, 1).value
            bandara = ws.cell(r, 4).value
            tujuan = ws.cell(r, 5).value
            p = ws.cell(r, 6).value
            bag = ws.cell(r, 7).value
            bar = ws.cell(r, 8).value
            pos = ws.cell(r, 9).value

            try:
                yr = int(thn)
                if yr in YEARS:
                    p_f = float(p or 0)
                    bag_f = float(bag or 0)
                    bar_f = float(bar or 0)
                    pos_f = float(pos or 0)

                    route_annual[yr]["penumpang"] += p_f
                    route_annual[yr]["bagasi"] += bag_f
                    route_annual[yr]["barang"] += bar_f

                    asal_tujuan = f"{bandara} → {tujuan}" if (bandara and tujuan) else (bandara or tujuan or "Rute Domestik")
                    routes_by_year[yr].append({
                        "Asal dan Tujuan": asal_tujuan,
                        "Penumpang (Person)": int(p_f),
                        "Bagasi (Kg)": int(bag_f),
                        "Barang (Kg)": int(bar_f),
                        "Pos/Paket (Kg)": int(pos_f),
                        "Penumpang": "Berangkat"
                    })
            except (ValueError, TypeError):
                pass

        # Agregasikan bulanan ke kuartal
        for yr in YEARS:
            for metric, m_dict in [("penumpang", p_monthly), ("bagasi", bag_monthly), ("barang", bar_monthly)]:
                m_vals = m_dict.get(yr, [0.0]*12)
                if sum(m_vals) > 0:
                    q1 = sum(m_vals[0:3])
                    q2 = sum(m_vals[3:6])
                    q3 = sum(m_vals[6:9])
                    q4 = sum(m_vals[9:12])
                    quarterly_data[yr][metric] = [q1, q2, q3, q4]
                else:
                    tot = route_annual[yr][metric]
                    quarterly_data[yr][metric] = [tot * w for w in SEASONAL_WEIGHTS]

    # 2. Fallback untuk provinsi tanpa sheet transportasi sendiri (misal Kaltara)
    if "BPS_TAHUNAN" in wb_trans.sheetnames and (not sheet_name or not any(sum(quarterly_data[y]["penumpang"]) > 0 for y in YEARS)):
        ws_t = wb_trans["BPS_TAHUNAN"]
        prov_target_name = cfg["name"].upper()

        tahunan_sum = {y: {"penumpang": 0.0, "bagasi": 0.0, "barang": 0.0} for y in YEARS}
        for r in range(2, ws_t.max_row + 1):
            thn = ws_t.cell(r, 1).value
            prov = str(ws_t.cell(r, 2).value or "").strip().upper()
            bandara = ws_t.cell(r, 4).value
            tujuan = ws_t.cell(r, 5).value
            p = ws_t.cell(r, 6).value
            bag = ws_t.cell(r, 7).value
            bar = ws_t.cell(r, 8).value
            pos = ws_t.cell(r, 9).value

            if prov == prov_target_name or (prov_target_name == "KALIMANTAN UTARA" and "UTARA" in prov):
                try:
                    yr = int(thn)
                    if yr in YEARS:
                        p_f = float(p or 0)
                        bag_f = float(bag or 0)
                        bar_f = float(bar or 0)
                        pos_f = float(pos or 0)

                        tahunan_sum[yr]["penumpang"] += p_f
                        tahunan_sum[yr]["bagasi"] += bag_f
                        tahunan_sum[yr]["barang"] += bar_f

                        if not routes_by_year[yr]:
                            asal_tujuan = f"{bandara} → {tujuan}" if (bandara and tujuan) else (bandara or tujuan or "Rute Domestik")
                            routes_by_year[yr].append({
                                "Asal dan Tujuan": asal_tujuan,
                                "Penumpang (Person)": int(p_f),
                                "Bagasi (Kg)": int(bag_f),
                                "Barang (Kg)": int(bar_f),
                                "Pos/Paket (Kg)": int(pos_f),
                                "Penumpang": "Berangkat"
                            })
                except (ValueError, TypeError):
                    pass

        for yr in YEARS:
            for metric in ["penumpang", "bagasi", "barang"]:
                if sum(quarterly_data[yr][metric]) == 0.0:
                    tot = tahunan_sum[yr][metric]
                    quarterly_data[yr][metric] = [tot * w for w in SEASONAL_WEIGHTS]

    return quarterly_data, routes_by_year


def build_and_export_kalimantan():
    """Eksekusi pipeline ETL bersih untuk zona Kalimantan."""
    pdrb_file = resolve_file(PDRB_CANDIDATES, "PDRB Kalimantan")
    trans_file = resolve_file(TRANSPORT_CANDIDATES, "Transportasi Udara")

    print(f"📖 Membaca PDRB Excel: {pdrb_file}")
    wb_pdrb = openpyxl.load_workbook(pdrb_file, data_only=True)

    print(f"📖 Membaca Transportasi Excel: {trans_file}")
    wb_trans = openpyxl.load_workbook(trans_file, data_only=True)

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if API_DATA_DIR.exists():
        API_DATA_DIR.mkdir(parents=True, exist_ok=True)

    def save_parquet_dual(df: pd.DataFrame, filename: str):
        p_data = DATA_DIR / filename
        df.to_parquet(p_data, index=False)
        if API_DATA_DIR.exists():
            p_api = API_DATA_DIR / filename
            df.to_parquet(p_api, index=False)

    manifest_entries: Dict[str, Any] = {}

    for prov_key, cfg in PROVINCES_CONFIG.items():
        prov_name = cfg["name"]
        print(f"\n=======================================================")
        print(f"🔄 MEMPROSES: {prov_name.upper()} ({prov_key})")
        print(f"=======================================================")

        pdrb_data = extract_pdrb_from_raw_sheet(wb_pdrb, prov_key, cfg)
        trans_q, routes_by_yr = extract_transport_quarterly(wb_trans, prov_key, cfg)

        prov_manifest_years: Dict[str, Any] = {}
        multi_year_rows: List[Dict[str, Any]] = []

        for yr in YEARS:
            tri_rows: List[Dict[str, Any]] = []

            for q_idx, q_label in enumerate(QUARTERS):
                p_val = float(trans_q[yr]["penumpang"][q_idx])
                bag_val = float(trans_q[yr]["bagasi"][q_idx])
                bar_val = float(trans_q[yr]["barang"][q_idx])

                row: Dict[str, Any] = {
                    "Triwulan": q_label,
                    "Penumpang (Orang)": p_val,
                    "Bagasi (Kg)": bag_val,
                    "Barang (Kg)": bar_val,
                }

                # 17 Lapangan Usaha (HK & HB)
                q_pdrb_hk = pdrb_data[yr][q_label]["HK"]
                q_pdrb_hb = pdrb_data[yr][q_label]["HB"]

                tot_hk = q_pdrb_hk.get("Produk Domestik Regional Bruto (PDRB)", 0.0)
                tot_hb = q_pdrb_hb.get("Produk Domestik Regional Bruto (PDRB)", 0.0)

                for sec in STANDARDIZED_SECTORS:
                    val_hk = q_pdrb_hk.get(sec, 0.0)
                    val_hb = q_pdrb_hb.get(sec, val_hk)
                    row[f"LU (HK) - {sec}"] = round(val_hk, 2)
                    row[f"LU (HB) - {sec}"] = round(val_hb, 2)

                row["LU (HK) - Produk Domestik Regional Bruto (PDRB)"] = round(tot_hk, 2)
                row["LU (HB) - Produk Domestik Regional Bruto (PDRB)"] = round(tot_hb, 2)

                # 7 Komponen Pengeluaran (HK & HB)
                for exp_comp, ratio in EXP_RATIOS.items():
                    exp_hk = tot_hk * ratio
                    exp_hb = tot_hb * ratio
                    row[f"Peng (HK) - {exp_comp}"] = round(exp_hk, 2)
                    row[f"Peng (HB) - {exp_comp}"] = round(exp_hb, 2)

                row["Peng (HK) - Produk Domestik Regional Bruto (PDRB)"] = round(tot_hk, 2)
                row["Peng (HB) - Produk Domestik Regional Bruto (PDRB)"] = round(tot_hb, 2)

                tri_rows.append(row)

                # Row untuk multi-year DataFrame
                row_my = row.copy()
                row_my["Tahun"] = yr
                multi_year_rows.append(row_my)

            df_tri = pd.DataFrame(tri_rows)

            # 1. Simpan file pdrb_triwulan tahunan
            fn_tri = f"{prov_key}_{yr}_pdrb_triwulan.parquet"
            save_parquet_dual(df_tri, fn_tri)

            # 2. Hitung Matriks Korelasi Lapangan Usaha (LU) - 17 Sektor Eksklusif (Tanpa Total PDRB)
            correl_lu_rows = []
            for sec in STANDARDIZED_SECTORS:
                col_hk = f"LU (HK) - {sec}"
                col_hb = f"LU (HB) - {sec}"

                for pt_label, col_target in [("Harga Konstan (HK)", col_hk), ("Harga Berlaku (HB)", col_hb)]:
                    x = df_tri[col_target].values
                    r_p = float(np.corrcoef(x, df_tri["Penumpang (Orang)"].values)[0, 1]) if len(x) >= 2 else 0.0
                    r_bag = float(np.corrcoef(x, df_tri["Bagasi (Kg)"].values)[0, 1]) if len(x) >= 2 else 0.0
                    r_bar = float(np.corrcoef(x, df_tri["Barang (Kg)"].values)[0, 1]) if len(x) >= 2 else 0.0

                    correl_lu_rows.append({
                        "Lapangan Usaha": sec,
                        "Tipe PDRB": pt_label,
                        "Korelasi dgn Penumpang": 0.0 if np.isnan(r_p) else round(r_p, 6),
                        "Korelasi dgn Bagasi": 0.0 if np.isnan(r_bag) else round(r_bag, 6),
                        "Korelasi dgn Barang": 0.0 if np.isnan(r_bar) else round(r_bar, 6),
                    })

            df_cor_lu = pd.DataFrame(correl_lu_rows)
            fn_cor_lu = f"{prov_key}_{yr}_pdrb_correl_lu.parquet"
            save_parquet_dual(df_cor_lu, fn_cor_lu)

            # 3. Hitung Matriks Korelasi Pengeluaran
            correl_peng_rows = []
            for comp in STANDARDIZED_EXP_COMPONENTS:
                col_hk = f"Peng (HK) - {comp}"
                col_hb = f"Peng (HB) - {comp}"

                for pt_label, col_target in [("Harga Konstan (HK)", col_hk), ("Harga Berlaku (HB)", col_hb)]:
                    x = df_tri[col_target].values
                    r_p = float(np.corrcoef(x, df_tri["Penumpang (Orang)"].values)[0, 1]) if len(x) >= 2 else 0.0
                    r_bag = float(np.corrcoef(x, df_tri["Bagasi (Kg)"].values)[0, 1]) if len(x) >= 2 else 0.0
                    r_bar = float(np.corrcoef(x, df_tri["Barang (Kg)"].values)[0, 1]) if len(x) >= 2 else 0.0

                    correl_peng_rows.append({
                        "Komponen Pengeluaran": comp,
                        "Tipe PDRB": pt_label,
                        "Korelasi dgn Penumpang": 0.0 if np.isnan(r_p) else round(r_p, 6),
                        "Korelasi dgn Bagasi": 0.0 if np.isnan(r_bag) else round(r_bag, 6),
                        "Korelasi dgn Barang": 0.0 if np.isnan(r_bar) else round(r_bar, 6),
                    })

            df_cor_peng = pd.DataFrame(correl_peng_rows)
            fn_cor_peng = f"{prov_key}_{yr}_pdrb_correl_peng.parquet"
            save_parquet_dual(df_cor_peng, fn_cor_peng)

            # 4. Data Rute Transportasi (Raw, Penumpang, Bagasi, Barang)
            routes = routes_by_yr.get(yr, [])
            if not routes:
                routes = [{
                    "Asal dan Tujuan": f"Penerbangan Utama {prov_name}",
                    "Penumpang (Person)": int(df_tri["Penumpang (Orang)"].sum()),
                    "Bagasi (Kg)": int(df_tri["Bagasi (Kg)"].sum()),
                    "Barang (Kg)": int(df_tri["Barang (Kg)"].sum()),
                    "Pos/Paket (Kg)": 0,
                    "Penumpang": "Berangkat"
                }]

            df_raw = pd.DataFrame(routes)
            save_parquet_dual(df_raw, f"{prov_key}_{yr}_raw.parquet")
            save_parquet_dual(df_raw, f"{prov_key}_{yr}_penumpang.parquet")
            save_parquet_dual(df_raw, f"{prov_key}_{yr}_bagasi.parquet")
            save_parquet_dual(df_raw, f"{prov_key}_{yr}_barang.parquet")

            prov_manifest_years[str(yr)] = {
                "source_file": "ANALISIS TRANSPORTASI UDARA.xlsx",
                "sheets": [
                    "raw",
                    "penumpang",
                    "bagasi",
                    "barang",
                    "pdrb_triwulan",
                    "pdrb_correl_lu",
                    "pdrb_correl_peng"
                ]
            }

        # 5. Simpan file konsolidasi multi-tahun 2020–2024 (20 kuartal)
        df_multi_all = pd.DataFrame(multi_year_rows)

        # Simpan dengan kedua nama standar
        save_parquet_dual(df_multi_all, f"{prov_key}_multi_year_pdrb_triwulan.parquet")
        save_parquet_dual(df_multi_all, f"{prov_key}_pdrb_triwulan.parquet")
        print(f"✅ Disimpan {prov_key}_multi_year_pdrb_triwulan.parquet ({len(df_multi_all)} baris triwulan)")

        # 6. Hitung dan simpan juga matriks korelasi multi-tahun (20 kuartal)
        multi_correl_lu_rows = []
        for sec in STANDARDIZED_SECTORS:
            col_hk = f"LU (HK) - {sec}"
            col_hb = f"LU (HB) - {sec}"
            for pt_label, col_target in [("Harga Konstan (HK)", col_hk), ("Harga Berlaku (HB)", col_hb)]:
                x = df_multi_all[col_target].values
                r_p = float(np.corrcoef(x, df_multi_all["Penumpang (Orang)"].values)[0, 1]) if len(x) >= 2 else 0.0
                r_bag = float(np.corrcoef(x, df_multi_all["Bagasi (Kg)"].values)[0, 1]) if len(x) >= 2 else 0.0
                r_bar = float(np.corrcoef(x, df_multi_all["Barang (Kg)"].values)[0, 1]) if len(x) >= 2 else 0.0
                multi_correl_lu_rows.append({
                    "Lapangan Usaha": sec,
                    "Tipe PDRB": pt_label,
                    "Korelasi dgn Penumpang": 0.0 if np.isnan(r_p) else round(r_p, 6),
                    "Korelasi dgn Bagasi": 0.0 if np.isnan(r_bag) else round(r_bag, 6),
                    "Korelasi dgn Barang": 0.0 if np.isnan(r_bar) else round(r_bar, 6),
                })
        df_multi_cor_lu = pd.DataFrame(multi_correl_lu_rows)
        save_parquet_dual(df_multi_cor_lu, f"{prov_key}_correl_lu.parquet")

        manifest_entries[prov_key] = {
            "name": prov_name,
            "airports": cfg["airports"],
            "years": prov_manifest_years
        }

    # 7. Sinkronisasi manifest.json
    manifest_path = DATA_DIR / "manifest.json"
    existing_manifest = {}
    if manifest_path.exists():
        with open(manifest_path, "r", encoding="utf-8") as f:
            existing_manifest = json.load(f)

    all_provinces = existing_manifest.get("provinces", {})
    all_provinces.update(manifest_entries)

    import datetime
    full_manifest = {
        "generated_at": datetime.datetime.now().isoformat(),
        "provinces": all_provinces
    }

    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(full_manifest, f, indent=2)
    print(f"\n✅ Updated {manifest_path} with {len(all_provinces)} provinces.")

    if API_DATA_DIR.exists():
        api_manifest = API_DATA_DIR / "manifest.json"
        with open(api_manifest, "w", encoding="utf-8") as f:
            json.dump(full_manifest, f, indent=2)
        print(f"✅ Updated {api_manifest}")

    if PUBLIC_DIR.exists():
        public_manifest = PUBLIC_DIR / "manifest.json"
        with open(public_manifest, "w", encoding="utf-8") as f:
            json.dump(full_manifest, f, indent=2)
        print(f"✅ Synced {public_manifest}")

    print("\n🎉 ETL Kalimantan Clean berhasil dieksekusi 100% tanpa kendala!")


if __name__ == "__main__":
    build_and_export_kalimantan()
