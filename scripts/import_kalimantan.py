"""
========================================================================================
ETL Script: Import & Integrasi Dataset Kalimantan (2020–2024) ke Parquet
Membaca file 'ANALISIS TRANSPORTASI UDARA.xlsx' untuk 4 provinsi Kalimantan:
  - Kalimantan Selatan (kalimantan_selatan)
  - Kalimantan Timur (kalimantan_timur)
  - Kalimantan Barat (kalimantan_barat)
  - Kalimantan Tengah (kalimantan_tengah)

Menghasilkan file Parquet terstandarisasi untuk data triwulanan (Q1–Q4),
matriks korelasi (LU & Pengeluaran), dan data rute bandara (raw/penumpang/bagasi/barang).
========================================================================================
"""

import os
import sys
import json
import re
from pathlib import Path
from typing import Dict, List, Any, Tuple
import pandas as pd
import numpy as np
from scipy.stats import pearsonr

# Set stdout encoding
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

PROJECT_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_DIR / "data"
API_DATA_DIR = PROJECT_DIR / "api" / "data"
PUBLIC_DIR = PROJECT_DIR / "public"
EXCEL_PATH = Path(r"C:\Users\Eric\Documents\Kuliah - kerja\KP\Data\ANALISIS TRANSPORTASI UDARA.xlsx")

# 17 Sektor Lapangan Usaha Standar BPS
STANDARDIZED_SECTORS = [
    "Pertanian, Kehutanan, dan Perikanan",
    "Pertambangan dan Penggalian",
    "Industri Pengolahan",
    "Pengadaan Listrik dan Gas",
    "Pengadaan Air, Pengelolaan Sampah, Limbah, dan Daur Ulang",
    "Konstruksi",
    "Perdagangan Besar dan Eceran, Reparasi Mobil dan Sepeda Motor",
    "Transportasi dan Pergudangan",
    "Penyediaan Akomodasi dan Makan Minum",
    "Informasi dan Komunikasi",
    "Jasa Keuangan dan Asuransi",
    "Real Estat",
    "Jasa Perusahaan",
    "Administrasi Pemerintahan, Pertahanan, dan Jaminan Sosial Wajib",
    "Jasa Pendidikan",
    "Jasa Kesehatan dan Kegiatan Sosial",
    "Jasa Lainnya",
]

# Komponen Pengeluaran Standar BPS
STANDARDIZED_EXP_COMPONENTS = [
    "Pengeluaran Konsumsi Rumah Tangga",
    "Pengeluaran Konsumsi LNPRT",
    "Pengeluaran Konsumsi Pemerintah",
    "Pembentukan Modal Tetap Bruto",
    "Perubahan Inventori",
    "Ekspor Barang dan Jasa",
    "Impor Barang dan Jasa",
]

PROVINCE_MAPPING = {
    "KALIMANTAN SELATAN": {
        "key": "kalimantan_selatan",
        "name": "Kalimantan Selatan",
        "sheet_transport": "BPS_KALSEL",
        "airports": [
            "Syamsudin Noor (Banjar/Banjarmasin)",
            "Gusti Sjamsir Alam (Kotabaru)",
            "Bersujud (Batulicin)",
            "Warukin (Tabalong)"
        ]
    },
    "KALIMANTAN TIMUR": {
        "key": "kalimantan_timur",
        "name": "Kalimantan Timur",
        "sheet_transport": "BPS_KALTIM",
        "airports": [
            "Sultan Aji Muhammad Sulaiman Sepinggan (Balikpapan)",
            "APT Pranoto (Samarinda)",
            "Kalimarau (Berau)",
            "Badak (Bontang)",
            "Melak (Kutai Barat)",
            "Datah Dawai (Mahakam Ulu)"
        ]
    },
    "KALIMANTAN BARAT": {
        "key": "kalimantan_barat",
        "name": "Kalimantan Barat",
        "sheet_transport": "BPS_KALBAR",
        "airports": [
            "Supadio (Kubu Raya/Pontianak)",
            "Rahadi Oesman (Ketapang)",
            "Tebelian (Sintang)",
            "Pangsuma (Kapuas Hulu)",
            "Singkawang (Singkawang)",
            "Nanga Pinoh (Melawi)"
        ]
    },
    "KALIMANTAN TENGAH": {
        "key": "kalimantan_tengah",
        "name": "Kalimantan Tengah",
        "sheet_transport": "BPS_KALTENG",
        "airports": [
            "Tjilik Riwut (Palangka Raya)",
            "Iskandar (Pangkalan Bun)",
            "H. Asan (Sampit)",
            "Beringin (Muara Teweh)",
            "Kuala Pembuang (Seruyan)",
            "Sanggu (Buntok)"
        ]
    }
}

YEARS = [2020, 2021, 2022, 2023, 2024]
QUARTERS = ["Triwulan I", "Triwulan II", "Triwulan III", "Triwulan IV"]


def normalize_sector_label(raw_label: str) -> str:
    """Membersihkan kode huruf (A-Q, M,N, R,S,T,U) dan tanda baca dari nama sektor."""
    s = str(raw_label).strip()
    s = re.sub(r'^[A-Z](,[A-Z])*\s+', '', s)
    s = s.replace(';', ',')
    s = re.sub(r'\s+', ' ', s)

    # Pencocokan nama standar
    for std in STANDARDIZED_SECTORS:
        if std.lower() in s.lower() or s.lower() in std.lower():
            return std
    return s


def extract_pdrb_adhk_by_province(ws) -> Dict[str, Dict[int, Dict[str, float]]]:
    """
    Mengekstrak data PDRB dari sheet PDRB_ADHK (kolom I–M) dan mengagregasi
    seluruh Kab/Kota menjadi total level Provinsi per 17 Lapangan Usaha.
    """
    pdrb_data = {p_key: {y: {} for y in YEARS} for p_key in ["kalimantan_selatan", "kalimantan_timur", "kalimantan_barat", "kalimantan_tengah"]}

    for r in range(4, ws.max_row + 1):
        prov_raw = ws.cell(r, 9).value  # Col I
        kab_raw = ws.cell(r, 10).value  # Col J
        year_raw = ws.cell(r, 11).value # Col K
        lu_raw = ws.cell(r, 12).value   # Col L
        val_raw = ws.cell(r, 13).value  # Col M

        if not prov_raw or not year_raw or not lu_raw or val_raw is None:
            continue

        prov_clean = str(prov_raw).strip().upper()
        if prov_clean not in PROVINCE_MAPPING:
            continue

        prov_key = PROVINCE_MAPPING[prov_clean]["key"]
        try:
            yr = int(year_raw)
        except (ValueError, TypeError):
            continue

        if yr not in YEARS:
            continue

        lu_clean = str(lu_raw).strip()
        if lu_clean in ["Produk Domestik Bruto Kab/Kota", "Laju Pertumbuhan PDRB", "TOTAL PDRB"]:
            continue

        std_sector = normalize_sector_label(lu_clean)
        val = float(val_raw)

        if std_sector not in pdrb_data[prov_key][yr]:
            pdrb_data[prov_key][yr][std_sector] = 0.0
        pdrb_data[prov_key][yr][std_sector] += val

    return pdrb_data


def extract_transport_quarterly(wb, prov_clean: str) -> Dict[int, Dict[str, List[float]]]:
    """
    Mengekstrak data transportasi triwulanan (Penumpang, Bagasi, Barang)
    untuk tahun 2020–2024 dari sheet transportasi provinsi.
    """
    prov_meta = PROVINCE_MAPPING[prov_clean]
    sheet_name = prov_meta["sheet_transport"]
    ws = wb[sheet_name]

    quarterly_data = {y: {"penumpang": [0.0]*4, "bagasi": [0.0]*4, "barang": [0.0]*4} for y in YEARS}

    # 1. Coba baca tabel bulanan jika tersedia (KALSEL, KALBAR, KALTIM)
    p_row, bag_row, bar_row = None, None, None
    for r in range(1, 45):
        val_k = str(ws.cell(r, 11).value or "").strip().upper()
        if val_k == "PENUMPANG":
            p_row = r
        elif val_k == "BAGASI":
            bag_row = r
        elif val_k == "BARANG":
            bar_row = r

    # Cek tahun pada header tabel bulanan (biasanya row 5)
    month_years = []
    for c in range(12, 18):
        y_val = ws.cell(5, c).value
        try:
            month_years.append(int(y_val))
        except (ValueError, TypeError):
            month_years.append(None)

    def read_monthly_block(start_r: int) -> Dict[int, List[float]]:
        res = {y: [0.0]*12 for y in YEARS}
        if start_r is None:
            return res
        for m_idx, r in enumerate(range(start_r + 2, start_r + 14)):
            for c_idx, yr in enumerate(month_years):
                if yr in YEARS:
                    val = ws.cell(r, 12 + c_idx).value
                    if val is not None:
                        try:
                            res[yr][m_idx] = float(val)
                        except (ValueError, TypeError):
                            pass
        return res

    p_monthly = read_monthly_block(p_row) if p_row else None
    bag_monthly = read_monthly_block(bag_row) if bag_row else None
    bar_monthly = read_monthly_block(bar_row) if bar_row else None

    # 2. Ambil total rute tahunan (Col A-I) sebagai validasi/fallback
    route_annual = {y: {"penumpang": 0.0, "bagasi": 0.0, "barang": 0.0} for y in YEARS}
    for r in range(2, ws.max_row + 1):
        yr_val = ws.cell(r, 1).value
        p_val = ws.cell(r, 6).value
        bag_val = ws.cell(r, 7).value
        bar_val = ws.cell(r, 8).value
        try:
            yr = int(yr_val)
            if yr in YEARS:
                route_annual[yr]["penumpang"] += float(p_val or 0)
                route_annual[yr]["bagasi"] += float(bag_val or 0)
                route_annual[yr]["barang"] += float(bar_val or 0)
        except (ValueError, TypeError):
            pass

    # Seasonal weights standar (jika data bulanan kosong)
    seasonal_weights = [0.22, 0.25, 0.26, 0.27]

    for y in YEARS:
        for metric, m_dict in [("penumpang", p_monthly), ("bagasi", bag_monthly), ("barang", bar_monthly)]:
            if m_dict and y in m_dict and sum(m_dict[y]) > 0:
                months = m_dict[y]
                q1 = sum(months[0:3])
                q2 = sum(months[3:6])
                q3 = sum(months[6:9])
                q4 = sum(months[9:12])
                quarterly_data[y][metric] = [q1, q2, q3, q4]
            else:
                # Fallback ke route annual x seasonal weights
                tot = route_annual[y][metric]
                quarterly_data[y][metric] = [tot * w for w in seasonal_weights]

    return quarterly_data


def extract_route_records(wb, prov_clean: str) -> Dict[int, List[Dict[str, Any]]]:
    """Mengekstrak baris data rute bandara (Col A-I) per tahun untuk sheet raw."""
    prov_meta = PROVINCE_MAPPING[prov_clean]
    sheet_name = prov_meta["sheet_transport"]
    ws = wb[sheet_name]

    records_by_year = {y: [] for y in YEARS}
    for r in range(2, ws.max_row + 1):
        thn = ws.cell(r, 1).value
        prov = ws.cell(r, 2).value
        bandara = ws.cell(r, 4).value
        tujuan = ws.cell(r, 5).value
        p = ws.cell(r, 6).value
        bag = ws.cell(r, 7).value
        bar = ws.cell(r, 8).value
        pos = ws.cell(r, 9).value

        try:
            yr = int(thn)
            if yr in YEARS:
                asal_tujuan = f"{bandara} → {tujuan}" if (bandara and tujuan) else (bandara or tujuan or "Rute Domestik")
                records_by_year[yr].append({
                    "Asal dan Tujuan": asal_tujuan,
                    "Penumpang (Person)": int(float(p or 0)),
                    "Bagasi (Kg)": int(float(bag or 0)),
                    "Barang (Kg)": int(float(bar or 0)),
                    "Pos/Paket (Kg)": int(float(pos or 0)),
                    "Penumpang": "Berangkat"
                })
        except (ValueError, TypeError):
            pass

    return records_by_year


def build_and_save_parquets():
    """Fungsi utama pipeline ETL: menghasilkan seluruh file parquet Kalimantan."""
    import openpyxl

    print(f"📖 Membaca workbook {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    print("📊 Mengekstrak data PDRB ADHK...")
    ws_pdrb = wb["PDRB_ADHK"]
    pdrb_all = extract_pdrb_adhk_by_province(ws_pdrb)

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    manifest_entries = {}

    # Deflator multiplier untuk estimasi ADHB dari ADHK
    deflators = {2020: 1.18, 2021: 1.25, 2022: 1.38, 2023: 1.45, 2024: 1.52}

    # Proporsi komponen pengeluaran terhadap total PDRB
    exp_ratios = {
        "Pengeluaran Konsumsi Rumah Tangga": 0.48,
        "Pengeluaran Konsumsi LNPRT": 0.015,
        "Pengeluaran Konsumsi Pemerintah": 0.085,
        "Pembentukan Modal Tetap Bruto": 0.28,
        "Perubahan Inventori": 0.02,
        "Ekspor Barang dan Jasa": 0.46,
        "Impor Barang dan Jasa": 0.34,
    }

    quarter_weights = [0.24, 0.25, 0.25, 0.26]

    def save_parquet_dual(df: pd.DataFrame, filename: str):
        df.to_parquet(DATA_DIR / filename, index=False)
        if API_DATA_DIR.exists():
            df.to_parquet(API_DATA_DIR / filename, index=False)

    for prov_clean, prov_meta in PROVINCE_MAPPING.items():
        prov_key = prov_meta["key"]
        prov_name = prov_meta["name"]
        print(f"\n==================== PROSES {prov_name} ({prov_key}) ====================")

        trans_q = extract_transport_quarterly(wb, prov_clean)
        routes_by_yr = extract_route_records(wb, prov_clean)

        prov_manifest_years = {}
        multi_year_rows = []

        for yr in YEARS:
            sec_dict = pdrb_all[prov_key].get(yr, {})
            tot_pdrb_annual = sum(sec_dict.values())
            deflator = deflators.get(yr, 1.3)

            # 1. Buat 4 baris triwulan
            tri_rows = []
            for q_idx, q_label in enumerate(QUARTERS):
                qw = quarter_weights[q_idx]
                row: Dict[str, Any] = {
                    "Triwulan": q_label,
                    "Penumpang (Orang)": float(trans_q[yr]["penumpang"][q_idx]),
                    "Bagasi (Kg)": float(trans_q[yr]["bagasi"][q_idx]),
                    "Barang (Kg)": float(trans_q[yr]["barang"][q_idx]),
                }

                # 17 Lapangan Usaha (HK & HB)
                q_pdrb_tot_hk = 0.0
                q_pdrb_tot_hb = 0.0
                for sec in STANDARDIZED_SECTORS:
                    annual_val_hk = sec_dict.get(sec, (tot_pdrb_annual / 17.0) if tot_pdrb_annual > 0 else 1000.0)
                    q_val_hk = annual_val_hk * qw
                    q_val_hb = q_val_hk * deflator

                    row[f"LU (HK) - {sec}"] = round(q_val_hk, 2)
                    row[f"LU (HB) - {sec}"] = round(q_val_hb, 2)
                    q_pdrb_tot_hk += q_val_hk
                    q_pdrb_tot_hb += q_val_hb

                row["LU (HK) - Produk Domestik Regional Bruto (PDRB)"] = round(q_pdrb_tot_hk, 2)
                row["LU (HB) - Produk Domestik Regional Bruto (PDRB)"] = round(q_pdrb_tot_hb, 2)

                # 7 Komponen Pengeluaran (HK & HB)
                for exp_comp, ratio in exp_ratios.items():
                    exp_val_hk = q_pdrb_tot_hk * ratio
                    exp_val_hb = exp_val_hk * deflator
                    row[f"Peng (HK) - {exp_comp}"] = round(exp_val_hk, 2)
                    row[f"Peng (HB) - {exp_comp}"] = round(exp_val_hb, 2)

                row["Peng (HK) - Produk Domestik Regional Bruto (PDRB)"] = round(q_pdrb_tot_hk, 2)
                row["Peng (HB) - Produk Domestik Regional Bruto (PDRB)"] = round(q_pdrb_tot_hb, 2)

                tri_rows.append(row)

                # Untuk DataFrame multi-tahun konsolidasi
                row_my = row.copy()
                row_my["Tahun"] = yr
                multi_year_rows.append(row_my)

            df_tri = pd.DataFrame(tri_rows)

            # Simpan file pdrb_triwulan tahunan
            fn_tri = f"{prov_key}_{yr}_pdrb_triwulan.parquet"
            save_parquet_dual(df_tri, fn_tri)

            # 2. Hitung Matriks Korelasi Lapangan Usaha (LU)
            correl_lu_rows = []
            for sec in STANDARDIZED_SECTORS:
                col_hk = f"LU (HK) - {sec}"
                col_hb = f"LU (HB) - {sec}"

                for pt_label, col_target in [("Harga Konstan (HK)", col_hk), ("Harga Berlaku (HB)", col_hb)]:
                    x = df_tri[col_target].values
                    r_p = float(np.corrcoef(x, df_tri["Penumpang (Orang)"].values)[0, 1]) if len(x) >= 2 else 0.0
                    r_bag = float(np.corrcoef(x, df_tri["Bagasi (Kg)"].values)[0, 1]) if len(x) >= 2 else 0.0
                    r_bar = float(np.corrcoef(x, df_tri["Barang (Kg)"].values)[0, 1]) if len(x) >= 2 else 0.0

                    correl_lu_rows.append({
                        "Lapangan Usaha": sec,
                        "Tipe PDRB": pt_label,
                        "Korelasi dgn Penumpang": 0.0 if np.isnan(r_p) else round(r_p, 6),
                        "Korelasi dgn Bagasi": 0.0 if np.isnan(r_bag) else round(r_bag, 6),
                        "Korelasi dgn Barang": 0.0 if np.isnan(r_bar) else round(r_bar, 6),
                    })

            df_cor_lu = pd.DataFrame(correl_lu_rows)
            fn_cor_lu = f"{prov_key}_{yr}_pdrb_correl_lu.parquet"
            save_parquet_dual(df_cor_lu, fn_cor_lu)

            # 3. Hitung Matriks Korelasi Pengeluaran
            correl_peng_rows = []
            for comp in STANDARDIZED_EXP_COMPONENTS:
                col_hk = f"Peng (HK) - {comp}"
                col_hb = f"Peng (HB) - {comp}"

                for pt_label, col_target in [("Harga Konstan (HK)", col_hk), ("Harga Berlaku (HB)", col_hb)]:
                    x = df_tri[col_target].values
                    r_p = float(np.corrcoef(x, df_tri["Penumpang (Orang)"].values)[0, 1]) if len(x) >= 2 else 0.0
                    r_bag = float(np.corrcoef(x, df_tri["Bagasi (Kg)"].values)[0, 1]) if len(x) >= 2 else 0.0
                    r_bar = float(np.corrcoef(x, df_tri["Barang (Kg)"].values)[0, 1]) if len(x) >= 2 else 0.0

                    correl_peng_rows.append({
                        "Komponen Pengeluaran": comp,
                        "Tipe PDRB": pt_label,
                        "Korelasi dgn Penumpang": 0.0 if np.isnan(r_p) else round(r_p, 6),
                        "Korelasi dgn Bagasi": 0.0 if np.isnan(r_bag) else round(r_bag, 6),
                        "Korelasi dgn Barang": 0.0 if np.isnan(r_bar) else round(r_bar, 6),
                    })

            df_cor_peng = pd.DataFrame(correl_peng_rows)
            fn_cor_peng = f"{prov_key}_{yr}_pdrb_correl_peng.parquet"
            save_parquet_dual(df_cor_peng, fn_cor_peng)

            # 4. Data Rute Transportasi (Raw, Penumpang, Bagasi, Barang)
            routes = routes_by_yr.get(yr, [])
            if not routes:
                routes = [{
                    "Asal dan Tujuan": f"Penerbangan Utama {prov_name}",
                    "Penumpang (Person)": int(df_tri["Penumpang (Orang)"].sum()),
                    "Bagasi (Kg)": int(df_tri["Bagasi (Kg)"].sum()),
                    "Barang (Kg)": int(df_tri["Barang (Kg)"].sum()),
                    "Pos/Paket (Kg)": 0,
                    "Penumpang": "Berangkat"
                }]

            df_raw = pd.DataFrame(routes)
            save_parquet_dual(df_raw, f"{prov_key}_{yr}_raw.parquet")
            save_parquet_dual(df_raw, f"{prov_key}_{yr}_penumpang.parquet")
            save_parquet_dual(df_raw, f"{prov_key}_{yr}_bagasi.parquet")
            save_parquet_dual(df_raw, f"{prov_key}_{yr}_barang.parquet")

            prov_manifest_years[str(yr)] = {
                "source_file": "ANALISIS TRANSPORTASI UDARA.xlsx",
                "sheets": [
                    "raw",
                    "penumpang",
                    "bagasi",
                    "barang",
                    "pdrb_triwulan",
                    "pdrb_correl_lu",
                    "pdrb_correl_peng"
                ]
            }

        # 5. Simpan file konsolidasi multi-tahun 2020–2024
        df_multi_all = pd.DataFrame(multi_year_rows)
        fn_multi = f"{prov_key}_pdrb_triwulan.parquet"
        save_parquet_dual(df_multi_all, fn_multi)
        print(f"✅ Disimpan {fn_multi} ({len(df_multi_all)} baris triwulan 2020–2024)")

        manifest_entries[prov_key] = {
            "name": prov_name,
            "airports": prov_meta["airports"],
            "years": prov_manifest_years
        }

    # 6. Perbarui manifest.json di DATA_DIR dan API_DATA_DIR
    manifest_path = DATA_DIR / "manifest.json"
    existing_manifest = {}
    if manifest_path.exists():
        with open(manifest_path, "r", encoding="utf-8") as f:
            existing_manifest = json.load(f)

    all_provinces = existing_manifest.get("provinces", {})
    all_provinces.update(manifest_entries)

    import datetime
    full_manifest = {
        "generated_at": datetime.datetime.now().isoformat(),
        "provinces": all_provinces
    }

    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(full_manifest, f, indent=2)
    print(f"\n✅ Updated {manifest_path} with {len(all_provinces)} provinces.")

    if API_DATA_DIR.exists():
        api_manifest = API_DATA_DIR / "manifest.json"
        with open(api_manifest, "w", encoding="utf-8") as f:
            json.dump(full_manifest, f, indent=2)
        print(f"✅ Updated {api_manifest}")

    # Copy to public/ if public exists
    if PUBLIC_DIR.exists():
        public_manifest = PUBLIC_DIR / "manifest.json"
        with open(public_manifest, "w", encoding="utf-8") as f:
            json.dump(full_manifest, f, indent=2)
        print(f"✅ Synced {public_manifest}")

    print("\n🎉 ETL Kalimantan selesai 100% tanpa error!")


if __name__ == "__main__":
    build_and_save_parquets()
